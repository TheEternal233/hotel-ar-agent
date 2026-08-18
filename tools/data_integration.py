"""M01: 数据准备与基础对接工具"""
import os, json
from datetime import datetime
from langchain.tools import tool
from tools import BASE_DIR, CONFIG_DIR
from tools.doc_parser import get_info


@tool
def data_integration(action: str, source_path: str = "") -> str:
    """统一数据入口：管理科目对照、信用规则、数据源适配。action可选: validate_environment(验证环境), list_accounts(科目对照), list_credit_rules(信用规则), check_data_format(格式检查)"""
    if action == "validate_environment":
        import sys
        ok = [f"Python {sys.version_info.major}.{sys.version_info.minor}"]
        for dep in ["langchain_openai", "langgraph", "openpyxl", "fastapi", "dotenv"]:
            try: __import__(dep.replace("-","_")); ok.append(f"  OK {dep}")
            except ImportError: ok.append(f"  MISS {dep}")
        env_path = os.path.join(BASE_DIR, ".env")
        ok.append(f"  {'OK' if os.path.exists(env_path) else 'MISS'} .env")
        for f in ["account_mapping.json", "credit_rules.json", "bank_fee_rates.json"]:
            fp = os.path.join(CONFIG_DIR, f)
            ok.append(f"  {'OK' if os.path.exists(fp) else 'MISS'} config/{f}")
        return "\n".join(ok)

    if action == "list_accounts":
        path = os.path.join(CONFIG_DIR, "account_mapping.json")
        with open(path, "r", encoding="utf-8") as f: cfg = json.load(f)
        lines = ["=== accounts ==="]
        for k,v in cfg.get("accounts",{}).items(): lines.append(f"  {k}: {v}")
        lines.append("\n=== commission ===")
        for k,v in cfg.get("commission_rules",{}).items(): lines.append(f"  {k}: {v['rate']*100}%")
        return "\n".join(lines)

    if action == "list_credit_rules":
        path = os.path.join(CONFIG_DIR, "credit_rules.json")
        with open(path, "r", encoding="utf-8") as f: cfg = json.load(f)
        lines = ["=== credit rules ==="]
        for r in cfg.get("rules",[]):
            lines.append(f"  {r['customer_type']}: limit={r.get('credit_limit','N/A')} days={r['credit_days']}")
        bp = cfg.get("bad_debt_policy",{})
        lines.append("\n=== bad debt ===")
        for b in bp.get("aging_brackets",[]):
            rate = bp.get("provision_rates",{}).get(b["days"],0)
            lines.append(f"  {b['label']}: {rate*100}%")
        return "\n".join(lines)

    if action == "check_data_format":
        if not source_path or not os.path.exists(source_path): return f"not found: {source_path}"
        ext = os.path.splitext(source_path)[1].lower()
        if ext in (".xlsx",".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
            try:
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
                rows = sum(1 for _ in ws.iter_rows(min_row=2))
                return f"Excel OK. Sheet={ws.title} cols={len(headers)} rows={rows}\nHeaders: {headers}"
            finally:
                wb.close()
        return f"unsupported: {ext}"
    return f"unknown: {action}"