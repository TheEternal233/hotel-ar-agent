"""M04: 携程佣金与付款通知工具"""
import os, json
from datetime import datetime
from langchain.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from tools import BASE_DIR, CONFIG_DIR
from tools.doc_parser import read_mapped, CTRIP_MAPPING

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")


@tool
def ctrip_commission(settlement_path: str) -> str:
    """携程佣金计算：读取携程结算单xlsx，自动计算佣金、标注差异分类。"""
    if not os.path.exists(settlement_path):
        return f"error: file not found {settlement_path}"
    cfg_path = os.path.join(CONFIG_DIR, "account_mapping.json")
    with open(cfg_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    comm = cfg.get("commission_rules", {}).get("携程", {})
    rate = comm.get("rate", 0.15)
    tax_div = comm.get("tax_divisor", 1.06)

    raw = read_mapped(settlement_path, CTRIP_MAPPING,
                      cast={"amount": float, "nights": int, "commission": float})

    records = []
    for r in raw:
        rid = str(r.get("order_id", ""))
        if not rid: continue
        amt = r.get("amount", 0) or 0
        nights = r.get("nights", 0) or 0
        pc = r.get("commission", 0) or 0
        cc = round(amt / tax_div * rate, 2) if amt > 0 else 0
        records.append({"id": rid, "amt": amt, "nights": nights, "pc": pc, "cc": cc, "diff": round(pc - cc, 2)})

    ta = sum(r["amt"] for r in records)
    tc = sum(r["cc"] for r in records)
    tn = sum(r["nights"] for r in records)
    dc = sum(1 for r in records if abs(r["diff"]) > 0.01)

    out_path = os.path.join(BASE_DIR, f"ctrip_comm_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Payment Notice"
    items = [("Hotel","K03"),("Nights",tn),("Amount",round(ta,2)),("Commission",round(tc,2)),("Net",round(ta-tc,2))]
    for i,(k,v) in enumerate(items,1):
        ws1.cell(row=i,column=1,value=k).font=Font(bold=True); ws1.cell(row=i,column=2,value=v)
    ws2 = wb.create_sheet("Details")
    hdrs=["Order","Amount","Nights","PlatformComm","CalcComm","Diff"]
    for j,h in enumerate(hdrs,1):
        c=ws2.cell(row=1,column=j,value=h);c.fill=HEADER_FILL;c.font=HEADER_FONT;c.border=THIN_BORDER
    for i,r in enumerate(records,2):
        for j,v in enumerate([r["id"],r["amt"],r["nights"],r["pc"],r["cc"],r["diff"]],1):
            c=ws2.cell(row=i,column=j,value=v);c.border=THIN_BORDER
            if j==6 and abs(v)>0.01:c.fill=RED_FILL
    wb.save(out_path);wb.close()
    return f"done: {out_path} orders={len(records)} nights={tn} amt={ta:,.2f} comm={tc:,.2f} net={ta-tc:,.2f} diffs={dc}"