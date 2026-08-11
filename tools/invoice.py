"""M08: 发票管理工具 — 生成开票清单/核对发票数据"""

import os
from datetime import datetime
from langchain.tools import tool
import openpyxl
from openpyxl.styles import Font

from tools.doc_parser import read_rezen, read_sheet
from enums.common_enum import HEADER_FILL, HEADER_FONT, THIN_BORDER, YELLOW_FILL

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUT_DIR, exist_ok=True)

# 发票税率配置
TAX_RATES = {"住宿": 0.06, "餐饮": 0.06, "SPA": 0.06, "会议": 0.06, "其他": 0.06}


@tool
def invoice_gen(receivable_path: str, invoice_type: str = "普通发票") -> str:
    """发票管理：从应收数据生成开票清单Excel，自动分类计算税额。

    Args:
        receivable_path: PMS应收数据文件路径
        invoice_type: 发票类型 (普通发票/专用发票)
    """
    if not os.path.exists(receivable_path):
        return f"错误：文件不存在: {receivable_path}"

    records = read_rezen(receivable_path)
    if not records:
        return "未读取到有效数据"

    # 按协议单位/客户分组
    invoice_groups = {}
    for r in records:
        corp = r.get("corp", "").strip() or r.get("name", "").strip() or "散客"
        if corp not in invoice_groups:
            invoice_groups[corp] = []
        invoice_groups[corp].append(r)

    # 生成开票清单
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUT_DIR, f"开票清单_{now}.xlsx")

    wb = openpyxl.Workbook()

    # Sheet 1: 开票汇总
    ws1 = wb.active
    ws1.title = "开票汇总"
    hdrs = ["客户名称", "笔数", "合计金额(含税)", "税额(6%)", "不含税金额", "发票类型"]
    for j, h in enumerate(hdrs, 1):
        c = ws1.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    grand_total = 0
    grand_tax = 0
    ri = 2
    for corp, items in sorted(invoice_groups.items(), key=lambda x: -len(x[1])):
        total = round(sum(r["amount"] for r in items), 2)
        tax = round(total * 0.06, 2)
        net = round(total - tax, 2)
        grand_total += total
        grand_tax += tax
        vals = [corp, len(items), total, tax, net, invoice_type]
        for j, v in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=j, value=v)
            c.border = THIN_BORDER
            if total > 10000:
                c.fill = YELLOW_FILL
        ri += 1

    # 合计行
    vals_sum = ["合计", sum(len(v) for v in invoice_groups.values()),
                round(grand_total, 2), round(grand_tax, 2),
                round(grand_total - grand_tax, 2), ""]
    for j, v in enumerate(vals_sum, 1):
        c = ws1.cell(row=ri, column=j, value=v)
        c.font = Font(bold=True)
        c.border = THIN_BORDER

    # Sheet 2: 开票明细
    ws2 = wb.create_sheet("开票明细")
    detail_hdrs = ["客户", "日期", "房号", "金额(含税)", "税额", "不含税", "类别", "备注"]
    for j, h in enumerate(detail_hdrs, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    ri = 2
    for corp, items in sorted(invoice_groups.items()):
        for r in items:
            amt = r["amount"]
            tax = round(amt * 0.06, 2)
            net = round(amt - tax, 2)
            date_str = r["date"].strftime("%Y-%m-%d") if r["date"] else ""
            vals = [corp, date_str, r["room"], amt, tax, net, "住宿", r.get("remark", "")]
            for j, v in enumerate(vals, 1):
                ws2.cell(row=ri, column=j, value=v).border = THIN_BORDER
            ri += 1

    wb.save(out_path)
    wb.close()

    return (
        f"开票清单已生成\n"
        f"报告: {out_path}\n"
        f"客户数: {len(invoice_groups)}\n"
        f"总金额(含税): {grand_total:,.2f}\n"
        f"总税额(6%): {grand_tax:,.2f}\n"
        f"不含税: {grand_total - grand_tax:,.2f}"
    )


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        print(invoice_gen.invoke({"receivable_path": sys.argv[1]}))
    else:
        data_dir = os.path.join(BASE_DIR, "data", "清远", "协议企业对账")
        files = [f for f in os.listdir(data_dir) if f.endswith(".xlsx")]
        if files:
            print(invoice_gen.invoke({"receivable_path": os.path.join(data_dir, files[0])}))