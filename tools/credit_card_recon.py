"""M07: 信用卡对账 — 按支付方式分组 + YFD预付卡独立通道

支持 PMS报表(按付款代码) + POS银行流水 + PMS应收后台 + YFD银行流水
"""

import os
from datetime import datetime, timedelta
from langchain.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from tools import BASE_DIR
from tools.doc_parser import read_sheet, read_rezen

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

# 付款代码→支付方式映射 (清远酒店 PMS)
PAYMENT_CODE_MAP = {
    "9001": "现金",
    "9002": "信用卡",
    "9003": "借记卡",
    "9004": "支付宝",
    "9005": "微信",
    "9006": "OTA卡",
    "9007": "预付卡",
    "9008": "挂账",
    "9101": "YFD支付宝",
    "9102": "YFD微信",
}


def _read_pms_report(path):
    """读取 PMS报表 格式: 账单号 桌台 付款代码 付款描述 金额 转账备注 操作员"""
    headers, rows = read_sheet(path)
    payment_groups = {}
    for r in rows:
        code = str(r.get("付款代码", "")).strip()
        # Skip summary rows where payment code is not a valid numeric code
        if not code.isdigit():
            continue
        desc = str(r.get("付款描述", "")).strip()
        amt_val = r.get("金额", 0)
        try:
            amount = float(amt_val) if amt_val is not None else 0
        except (ValueError, TypeError):
            continue
        pay_type = PAYMENT_CODE_MAP.get(code, code)
        if pay_type not in payment_groups:
            payment_groups[pay_type] = {"count": 0, "total": 0, "records": []}
        payment_groups[pay_type]["count"] += 1
        payment_groups[pay_type]["total"] += amount
        payment_groups[pay_type]["records"].append(r)
    return payment_groups


def _read_pos_statement(path):
    """读取 POS机银行流水 格式 (含手续费)"""
    headers, rows = read_sheet(path)
    records = []
    for r in rows:
        amt = float(r.get("客户实付", r.get("交易金额", r.get("金额", 0))) or 0)
        fee = float(r.get("手续费", 0) or 0)
        net = float(r.get("入账金额", 0) or 0)
        pay_type = str(r.get("支付类型", "")).strip()
        records.append({"amount": amt, "fee": fee, "net": net, "pay_type": pay_type})
    return records


def _read_pms_ar_backend(path):
    """读取 PMS应收后台 格式 (rezen-like with additional fields)"""
    return read_rezen(path)


def batch_card_recon(data_dir=None):
    """批量信用卡对账：自动配对 PMS报表 + PMS应收后台 + 银行流水"""
    if data_dir is None:
        data_dir = os.path.join(BASE_DIR, "data", "清远", "信用卡对账")
    if not os.path.exists(data_dir):
        return f"错误：数据目录不存在: {data_dir}"

    files = [f for f in os.listdir(data_dir) if f.endswith(".xlsx")]

    # 分类文件
    pms_report_file = None
    pos_file = None
    wechat_ar = None
    alipay_ar = None
    yfd_wechat_ar = None
    yfd_alipay_ar = None
    yfd_wechat_bank = None
    yfd_alipay_bank = None

    for f in files:
        fl = f.lower()
        if "pms报表" in fl or "pms报表" in f:
            pms_report_file = f
        elif "pos" in fl or "pos机" in fl or "银行流水" in fl and "yfd" not in fl:
            pos_file = f
        elif "wechat" in fl and "应收后台" in f and "yfd" not in fl:
            wechat_ar = f
        elif "alipay" in fl and "应收后台" in f and "yfd" not in fl:
            alipay_ar = f
        elif "yfd" in fl and "wechat" in fl and "应收后台" in f:
            yfd_wechat_ar = f
        elif "yfd" in fl and "alipay" in fl and "应收后台" in f:
            yfd_alipay_ar = f
        elif "yfd" in fl and "wechat" in fl and "银行流水" in f:
            yfd_wechat_bank = f
        elif "yfd" in fl and "alipay" in fl and "银行流水" in f:
            yfd_alipay_bank = f

    results = []
    out_dir = os.path.join(BASE_DIR, "output")
    os.makedirs(out_dir, exist_ok=True)

    # 1. PMS报表 支付方式分组
    if pms_report_file:
        path = os.path.join(data_dir, pms_report_file)
        try:
            groups = _read_pms_report(path)
            pms_summary = {k: {"count": v["count"], "total": round(v["total"], 2)}
                          for k, v in groups.items()}
            results.append(f"PMS报表: {len(groups)}种支付方式, 总记录{sum(v['count'] for v in groups.values())}条")
            for pay_type, info in pms_summary.items():
                results.append(f"  {pay_type}: {info['count']}笔  {info['total']:,.2f}")
        except Exception as e:
            results.append(f"PMS报表读取失败: {e}")

    # 2. POS银行流水
    if pos_file:
        path = os.path.join(data_dir, pos_file)
        try:
            pos_records = _read_pos_statement(path)
            pos_total = sum(r["amount"] for r in pos_records)
            pos_fee = sum(r["fee"] for r in pos_records)
            results.append(f"\nPOS流水: {len(pos_records)}笔 金额{pos_total:,.2f} 手续费{pos_fee:,.2f}")
        except Exception as e:
            results.append(f"\nPOS流水读取失败: {e}")

    # 3. 微信PMS应收后台
    if wechat_ar:
        path = os.path.join(data_dir, wechat_ar)
        try:
            records = _read_pms_ar_backend(path)
            total = sum(r["amount"] for r in records)
            results.append(f"\n微信PMS应收: {len(records)}笔 金额{total:,.2f}")
        except Exception as e:
            results.append(f"\n微信PMS应收读取失败: {e}")

    # 4. 支付宝PMS应收后台
    if alipay_ar:
        path = os.path.join(data_dir, alipay_ar)
        try:
            records = _read_pms_ar_backend(path)
            total = sum(r["amount"] for r in records)
            results.append(f"支付宝PMS应收: {len(records)}笔 金额{total:,.2f}")
        except Exception as e:
            results.append(f"支付宝PMS应收读取失败: {e}")

    # 5. YFD预付卡 (微信)
    yfd_results = []
    if yfd_wechat_ar:
        path_ar = os.path.join(data_dir, yfd_wechat_ar)
        try:
            records = _read_pms_ar_backend(path_ar)
            total = sum(r["amount"] for r in records)
            yfd_results.append(f"YFD微信PMS应收: {len(records)}笔 金额{total:,.2f}")
        except Exception as e:
            yfd_results.append(f"YFD微信PMS应收读取失败: {e}")

    if yfd_wechat_bank:
        path_bank = os.path.join(data_dir, yfd_wechat_bank)
        try:
            headers, rows = read_sheet(path_bank)
            yfd_results.append(f"YFD微信银行流水: {len(rows)}行")
        except Exception as e:
            yfd_results.append(f"YFD微信银行流水读取失败: {e}")

    if yfd_alipay_ar:
        path_ar = os.path.join(data_dir, yfd_alipay_ar)
        try:
            records = _read_pms_ar_backend(path_ar)
            total = sum(r["amount"] for r in records)
            yfd_results.append(f"YFD支付宝PMS应收: {len(records)}笔 金额{total:,.2f}")
        except Exception as e:
            yfd_results.append(f"YFD支付宝PMS应收读取失败: {e}")

    if yfd_alipay_bank:
        path_bank = os.path.join(data_dir, yfd_alipay_bank)
        try:
            headers, rows = read_sheet(path_bank)
            yfd_results.append(f"YFD支付宝银行流水: {len(rows)}行")
        except Exception as e:
            yfd_results.append(f"YFD支付宝银行流水读取失败: {e}")

    if yfd_results:
        results.append("\n[YFD预付卡]")
        results.extend(yfd_results)

    # 生成汇总报告
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"信用卡对账汇总_{now}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "对账汇总"
    for i, line in enumerate(results, 1):
        ws.cell(row=i, column=1, value=line)

    wb.save(out_path)
    wb.close()

    return "信用卡批量对账完成\n" + "\n".join(results) + f"\n\n报告: {out_path}"


@tool
def credit_card_recon(bank_statement_path: str = "", pms_card_path: str = "") -> str:
    """信用卡对账：支持批量模式(无参数)和单通道模式。
    批量模式自动读取 data/清远/信用卡对账 目录下所有文件并生成汇总。
    单通道模式按金额+日期+卡号三维匹配，生成差异分类表。
    """
    if not bank_statement_path and not pms_card_path:
        return batch_card_recon()

    # 单通道模式 (原有逻辑)
    for p in [bank_statement_path, pms_card_path]:
        if not os.path.exists(p):
            return f"error: {p}"

    def _parse_date(val):
        if isinstance(val, datetime):
            return val
        if isinstance(val, str) and val.strip():
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"]:
                try:
                    return datetime.strptime(val.strip(), fmt)
                except ValueError:
                    pass
        return None

    def _read_card(path):
        headers, rows = read_sheet(path)
        records = []
        for r in rows:
            date_val = r.get("日期", r.get("交易日期", r.get("date", "")))
            amount_val = r.get("金额", r.get("交易金额", r.get("amount", 0)))
            card_val = r.get("卡号", r.get("card", ""))
            records.append({
                "date": _parse_date(date_val),
                "amount": float(amount_val or 0),
                "card": str(card_val).strip()[-4:],
                "raw": r
            })
        return records

    bank = _read_card(bank_statement_path)
    pms = _read_card(pms_card_path)

    matched_bank = set()
    matched_pms = set()
    match_results = []
    for bi, br in enumerate(bank):
        for pi, pr in enumerate(pms):
            if pi in matched_pms:
                continue
            if abs(br["amount"] - pr["amount"]) > 0.01:
                continue
            if br["date"] and pr["date"] and abs((br["date"] - pr["date"]).days) > 1:
                continue
            if br["card"] and pr["card"] and br["card"] != pr["card"]:
                continue
            matched_bank.add(bi)
            matched_pms.add(pi)
            match_results.append({"bank": br, "pms": pr, "status": "match"})
            break

    unmatched_bank = [br for i, br in enumerate(bank) if i not in matched_bank]
    unmatched_pms = [pr for i, pr in enumerate(pms) if i not in matched_pms]
    shortfall = sum(r["amount"] for r in unmatched_bank)
    overage = sum(r["amount"] for r in unmatched_pms)
    total = sum(r["amount"] for r in bank)
    fee_rate = 0.01
    fee = round(total * fee_rate, 2)

    out_dir = os.path.join(BASE_DIR, "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"card_recon_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "balance"
    items = [
        ("bank_balance", round(total, 2)), ("+overage", round(overage, 2)),
        ("-shortfall", round(shortfall, 2)), ("", ""),
        ("adjusted", round(total + overage - shortfall, 2)),
        ("", ""), (f"fee({fee_rate*100}%)", fee), ("net", round(total - fee, 2))
    ]
    for i, (k, v) in enumerate(items, 1):
        ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=v)

    ws2 = wb.create_sheet("diffs")
    hdrs = ["type", "date", "amount", "card", "note"]
    for j, h in enumerate(hdrs, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER
    ri = 2
    for r in unmatched_bank:
        vals = ["shortfall", str(r["date"].date()) if r["date"] else "", r["amount"], r["card"], "PMS missing"]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=j, value=v)
            c.fill = RED_FILL
            c.border = THIN_BORDER
        ri += 1
    for r in unmatched_pms:
        vals = ["overage", str(r["date"].date()) if r["date"] else "", r["amount"], r["card"], "bank not received"]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=j, value=v)
            c.fill = YELLOW_FILL
            c.border = THIN_BORDER
        ri += 1

    ws3 = wb.create_sheet("voucher")
    vhdrs = ["date", "summary", "account", "name", "debit", "credit"]
    for j, h in enumerate(vhdrs, 1):
        c = ws3.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER
    voucher = [
        [datetime.now().strftime("%Y-%m-%d"), "card fee", "6602", "fee expense", fee, ""],
        [datetime.now().strftime("%Y-%m-%d"), "card fee", "1002", "bank deposit", "", fee],
    ]
    for i, row in enumerate(voucher, 2):
        for j, v in enumerate(row, 1):
            ws3.cell(row=i, column=j, value=v).border = THIN_BORDER
    wb.save(out_path)
    wb.close()

    return (
        f"card recon done: {out_path}\n"
        f"bank={len(bank)} {total:,.2f}\n"
        f"match={len(match_results)}\n"
        f"shortfall={len(unmatched_bank)} {shortfall:,.2f}\n"
        f"overage={len(unmatched_pms)} {overage:,.2f}\n"
        f"fee={fee:.2f}"
    )


if __name__ == "__main__":
    result = batch_card_recon()
    print(result)
