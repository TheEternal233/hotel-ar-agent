"""M03: 账龄分析与坏账计提工具"""
import os, json
from datetime import datetime, timedelta
from langchain.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from tools.doc_parser import read_mapped, AGING_MAPPING, _parse_date
from tools import BASE_DIR, CONFIG_DIR

RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


def _load_credit_rules():
    path = os.path.join(CONFIG_DIR, "credit_rules.json")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# _parse_date moved to doc_parser


@tool
def aging_analysis(receivable_path: str, as_of_date: str = "") -> str:
    """账龄分析与坏账计提：读取应收台账xlsx，按逾期天数分段统计账龄，自动计提坏账准备，生成账龄分析报表和坏账凭证。"""
    """账龄分析与坏账计提：读取应收台账xlsx，按逾期天数分段统计账龄，自动计提坏账准备，生成账龄分析报表和坏账凭证。"""
    if not os.path.exists(receivable_path):
        return f"错误：文件不存在 {receivable_path}"

    rules = _load_credit_rules()
    provision_rates = rules["bad_debt_policy"]["provision_rates"]

    if not as_of_date:
        as_of_date = datetime.now()
    else:
        as_of_date = _parse_date(as_of_date) or datetime.now()

    # 使用标准 doc_parser 读取 — 替代原有 openpyxl + col_map 逻辑
    raw = read_mapped(receivable_path, AGING_MAPPING,
                      cast={"due_date": _parse_date, "amount": float})
    records = []
    for r in raw:
        due = r.get("due_date")
        amt = r.get("amount")
        if due is None or amt is None:
            continue
        overdue_days = (as_of_date - due).days
        records.append({
            "customer": str(r.get("customer", "")),
            "due_date": due,
            "amount": amt,
            "overdue_days": max(0, overdue_days),
        })

    bracket_defs = [
        (range(1, 31), "1-30"), (range(31, 61), "31-60"),
        (range(61, 91), "61-90"), (range(91, 121), "91-120"),
        (range(121, 181), "121-180"), (range(181, 9999), "180+"),
    ]

    summary = {}
    for r in records:
        cust = r["customer"]
        if cust not in summary:
            summary[cust] = {"amounts": {k: 0 for _, k in bracket_defs}, "total": 0}
        for brange, bkey in bracket_defs:
            if r["overdue_days"] in brange:
                summary[cust]["amounts"][bkey] += r["amount"]
                break
        summary[cust]["total"] += r["amount"]

    out_path = os.path.join(BASE_DIR, f"账龄分析报表_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "账龄分析"
    hdrs1 = ["客户名称", "1-30天", "31-60天", "61-90天", "91-120天", "121-180天", "180天以上", "合计", "坏账准备"]
    for j, h in enumerate(hdrs1, 1):
        c = ws1.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    total_provision = 0
    ri = 2
    for cust, data in sorted(summary.items(), key=lambda x: -x[1]["total"]):
        provision = sum(data["amounts"][k] * provision_rates.get(k, 0) for k in data["amounts"])
        total_provision += provision
        vals = [cust] + [data["amounts"][k] for _, k in bracket_defs] + [data["total"], round(provision, 2)]
        for j, v in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=j, value=v)
            c.border = THIN_BORDER
            if j >= 6 and isinstance(v, (int, float)) and v > 0:
                c.fill = RED_FILL
        ri += 1

    total_row = ["合计"] + [sum(summary[c]["amounts"][k] for c in summary) for _, k in bracket_defs]
    total_row += [sum(total_row[1:-1]), round(total_provision, 2)]
    for j, v in enumerate(total_row, 1):
        c = ws1.cell(row=ri, column=j, value=v)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        c.border = THIN_BORDER

    ws2 = wb.create_sheet("坏账凭证")
    vhdrs = ["凭证日期", "摘要", "科目代码", "科目名称", "借方金额", "贷方金额"]
    for j, h in enumerate(vhdrs, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER
    voucher_rows = [
        [as_of_date.strftime("%Y-%m-%d"), "计提坏账准备", "6701", "资产减值损失", round(total_provision, 2), ""],
        [as_of_date.strftime("%Y-%m-%d"), "计提坏账准备", "1231", "坏账准备", "", round(total_provision, 2)],
    ]
    for i, row in enumerate(voucher_rows, 2):
        for j, v in enumerate(row, 1):
            ws2.cell(row=i, column=j, value=v).border = THIN_BORDER

    wb.save(out_path)
    wb.close()

    bracket_summary = "\n".join(f"  {k}: {summary.get(k, {}).get('total', 0):,.2f}" for k in summary)
    return (
        f"账龄分析完成: {out_path}\n"
        f"截止日期: {as_of_date.strftime('%Y-%m-%d')}\n"
        f"客户数: {len(summary)}\n"
        f"坏账准备: {total_provision:,.2f}\n\n"
        f"账龄分布:\n{bracket_summary}"
    )
