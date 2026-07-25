"""M05: 每日单据核对工具"""
import os
from datetime import datetime
from langchain.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from tools import BASE_DIR
from tools.doc_parser import read_sheet

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

SMALL = 0.10
WARN = 100.00

@tool
def daily_night_audit_check(night_report_path: str, ar_ledger_path: str) -> str:
    """每日夜审报表核对：读取夜审报表和应收台账xlsx，逐笔比对挂账与收款，差异分级告警，生成日结报告。"""
    for p in [night_report_path, ar_ledger_path]:
        if not os.path.exists(p): return f"error: {p}"

    def _build_index(path):
        headers, rows = read_sheet(path)
        data = {}
        for row in rows:
            room = str(row.get("room", row.get("房号", "")))
            acct = str(row.get("account", row.get("账号", "")))
            date = str(row.get("date", row.get("日期", "")))
            key = f"{room}|{acct}|{date}"
            data[key] = row
        return data

    night = _build_index(night_report_path)
    ar = _build_index(ar_ledger_path)
    all_keys = sorted(set(night) | set(ar))
    stats = {"total": 0, "match": 0, "small": 0, "alert": 0, "night_only": 0, "ar_only": 0}

    results = []
    for key in all_keys:
        n = night.get(key); a = ar.get(key)
        if n and a:
            stats["total"] += 1
            namt = float(n.get("amount", 0) or 0)
            aamt = float(a.get("amount", 0) or 0)
            diff = abs(namt - aamt)
            if diff <= SMALL: stats["match"] += 1; level = "match"
            elif diff <= WARN: stats["small"] += 1; level = "warn"
            else: stats["alert"] += 1; level = "alert"
            results.append({"key": key, "n": n, "a": a, "diff": round(namt - aamt, 2), "level": level})
        elif n: stats["total"] += 1; stats["night_only"] += 1
        else: stats["total"] += 1; stats["ar_only"] += 1

    out_path = os.path.join(BASE_DIR, f"daily_check_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb = Workbook()
    ws1 = wb.active; ws1.title = "summary"
    items = [("date", datetime.now().strftime("%Y-%m-%d")), ("total", stats["total"]), ("match", stats["match"]), ("warn", stats["small"]), ("alert", stats["alert"])]
    for i,(k,v) in enumerate(items,1):
        ws1.cell(row=i,column=1,value=k).font=Font(bold=True)
        c = ws1.cell(row=i,column=2,value=v)
        if k == "alert": c.fill = RED_FILL; c.font = Font(bold=True, color="FF0000")

    ws2 = wb.create_sheet("alerts")
    hdrs = ["room", "account", "date", "night", "ar", "diff", "level"]
    for j,h in enumerate(hdrs,1):
        c=ws2.cell(row=1,column=j,value=h);c.fill=HEADER_FILL;c.font=HEADER_FONT;c.border=THIN_BORDER
    ri = 2
    for r in results:
        if r["level"] == "match": continue
        parts = r["key"].split("|")
        namt = float((r["n"] or {}).get("amount", 0) or 0)
        aamt = float((r["a"] or {}).get("amount", 0) or 0)
        vals = [parts[0] if len(parts)>0 else "", parts[1] if len(parts)>1 else "", parts[2] if len(parts)>2 else "", namt, aamt, r["diff"] or "", r["level"]]
        for j,v in enumerate(vals,1):
            c=ws2.cell(row=ri,column=j,value=v);c.border=THIN_BORDER
            if r["level"] == "alert": c.fill = RED_FILL
            elif r["level"] in ("warn",): c.fill = YELLOW_FILL
        ri += 1
    wb.save(out_path); wb.close()
    return f"daily check done: {out_path} total={stats['total']} match={stats['match']} warn={stats['small']} alert={stats['alert']}"
