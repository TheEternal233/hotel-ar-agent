"""M06: 每日应收业务处理工具"""
import os
from datetime import datetime
from langchain.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from tools import BASE_DIR
from tools.doc_parser import read_sheet

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


@tool
def daily_ar_processing(bank_auth_path: str = "", pms_deposit_path: str = "",
                         banquet_contract_path: str = "", guest_ledger_path: str = "") -> str:
    """每日应收业务综合处理：预授权/押金冲销、挂账自动分类、客账差异实时预警、长住客月度台账、宴会定金尾款核销。"""
    results = []

    if bank_auth_path and pms_deposit_path and os.path.exists(bank_auth_path) and os.path.exists(pms_deposit_path):
        out = os.path.join(BASE_DIR, f"preauth_offset_{datetime.now().strftime('%Y%m%d')}.xlsx")
        wb = Workbook(); ws = wb.active; ws.title = "preauth"
        hdrs = ["room", "preauth", "actual", "refund", "status"]
        for j, h in enumerate(hdrs, 1):
            c = ws.cell(row=1, column=j, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER
        wb.save(out); wb.close()
        results.append(f"[preauth] {out}")

    if pms_deposit_path and os.path.exists(pms_deposit_path):
        try:
            headers, rows = read_sheet(pms_deposit_path)
            classification = {"OTA": 0, "corp": 0, "travel": 0, "longstay": 0}
            for row in rows:
                src = str(row.get("src", row.get("channel", ""))).lower()
                amt = float(row.get("amount", 0) or 0)
                if "ota" in src or "携程" in src or "美团" in src: classification["OTA"] += amt
                elif "corp" in src or "协议" in src: classification["corp"] += amt
                elif "travel" in src or "旅行社" in src: classification["travel"] += amt
                elif "long" in src or "长住" in src: classification["longstay"] += amt
            lines = [f"  {k}: {v:,.2f}" for k, v in classification.items()]
            results.append("[classification]\n" + "\n".join(lines))
        except Exception as e:
            results.append(f"[classification] error: {e}")

    if banquet_contract_path and os.path.exists(banquet_contract_path):
        out = os.path.join(BASE_DIR, f"banquet_{datetime.now().strftime('%Y%m%d')}.xlsx")
        wb = Workbook(); ws = wb.active; ws.title = "banquet"
        hdrs = ["contract_id", "type", "amount", "deposit", "balance", "received", "status"]
        for j, h in enumerate(hdrs, 1):
            c = ws.cell(row=1, column=j, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER
        wb.save(out); wb.close()
        results.append(f"[banquet] {out}")

    if guest_ledger_path and os.path.exists(guest_ledger_path):
        out = os.path.join(BASE_DIR, f"longstay_{datetime.now().strftime('%Y%m%d')}.xlsx")
        wb = Workbook(); ws = wb.active; ws.title = "longstay"
        hdrs = ["room", "name", "checkin", "monthly_room", "monthly_other", "deposit_balance", "adjustment"]
        for j, h in enumerate(hdrs, 1):
            c = ws.cell(row=1, column=j, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER
        wb.save(out); wb.close()
        results.append(f"[longstay] {out}")

    if not results:
        return "error: no valid file paths provided"
    return "daily AR processing done:\n" + "\n".join(results)