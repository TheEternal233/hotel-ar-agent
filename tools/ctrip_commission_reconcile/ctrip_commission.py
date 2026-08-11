# -*- coding: utf-8 -*-
"""M04: 携程佣金核对工具（整合 Parser + Reconciler）"""
import logging
import os
from datetime import datetime
from typing import Optional

import pandas as pd
from langchain_core.tools import tool

from tools.ctrip_commission_reconcile.CommissionReconciler import CommissionReconciler
from tools.ctrip_commission_reconcile.CtripCommissionParser import CtripCommissionParser
from tools.ctrip_commission_reconcile.PmsParser import PmsParser
from openpyxl.styles import PatternFill, Alignment

from enums.common_enum import HEADER_FILL, HEADER_FONT, THIN_BORDER

logger=logging.getLogger(__name__)
# 状态颜色映射
_STATUS_FILL = {
    "OK":           PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # 淡绿
    "DIFF":         PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # 淡红
    "UNMATCHED":    PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # 淡黄
    "NO_PMS_PRICE": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # 淡黄
    "PMS_ONLY": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),      # 淡蓝
}
# 项目根目录（tools/ 的上一级）
_BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
UPLOAD_DIR = os.path.join(_BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(_BASE_DIR, "output")


def _apply_header_style(ws):
    """给表头应用统一样式"""
    for cell in ws[1]:
        cell.fill=HEADER_FILL
        cell.font=HEADER_FONT
        cell.border=THIN_BORDER
        cell.alignment=Alignment(horizontal="center", vertical="center")

def _find_status_col_idx(df:pd.DataFrame)->Optional[int]:
    """找到status列的1-based索引"""
    for idx,col_name in enumerate(df.columns,start=1):
        if col_name=="status":
            return idx
    return None


def _apply_row_style(ws,min_row:int,max_row:int,status_col_idx:Optional[int]=None):
    """给数据行加边框，并按status染色"""
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        status_val=row[status_col_idx-1].value if status_col_idx else None
        fill=_STATUS_FILL.get(status_val)
        for cell in row:
            cell.border=THIN_BORDER
            cell.alignment=Alignment(vertical="center")
            if fill:
                cell.fill=fill




@tool
def ctrip_commission(ctrip_filename:str,pms_filename:Optional[str]=None)->str:
    """
    携程佣金核对：读取 uploads 下的携程结算单，可选加载 PMS，自动匹配验算后输出差异表到 output。

    Args:
        ctrip_filename: uploads 目录下的携程文件名，如 "携程佣金.xls"
        pms_filename: uploads 目录下的 PMS 文件名（可选），如 "pms.xlsx"
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 读取携程
    ctrip_path=os.path.join(UPLOAD_DIR,ctrip_filename)
    if not os.path.exists(ctrip_path):
        return f"error: file not found: {ctrip_path}"

    ctrip_parser = CtripCommissionParser()
    try:
        ctrip_records = ctrip_parser.parse(ctrip_path)
    except ValueError as e:
        return f"error:携程账单解析失败 - {e}"

    if not ctrip_records:
        return f"error: 携程账单解析为空，请检查文件内容"

    total_orders=len(ctrip_records)
    total_nights=sum(r["nights"] for r in ctrip_records)
    total_comm=sum(r["commission"] for r in ctrip_records)

    #读取PMS
    has_pms=bool(pms_filename and pms_filename.strip())
    pms_records=[]

    if has_pms:
        pms_path=os.path.join(UPLOAD_DIR,pms_filename)
        if not os.path.exists(pms_path):
            return f"error: PMS file not found: {pms_path}"
        pms_parser = PmsParser()

        try:
            pms_records = pms_parser.parse(pms_path)
        except ValueError as e:
            return f"error:PMS账单解析失败 - {e}"

    # 核对
    recon=CommissionReconciler(commission_rate=0.15,tolerance=0.05)
    recon.load_ctrip(ctrip_records)

    if pms_records:
        recon.load_pms(pms_records)
    result=recon.run()
    stats=recon.summary()



    # 输出Excel
    timestamp=datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name=f"ctrip_recon_{timestamp}.xlsx"
    out_path=os.path.join(OUTPUT_DIR,out_name)

    df=pd.DataFrame(result)
    status_col_idx=_find_status_col_idx(df)
    with pd.ExcelWriter(out_path,engine="openpyxl") as writer:
        #sheet1 全部结果
        df.to_excel(writer,sheet_name="核对结果",index=False)
        ws=writer.sheets["核对结果"]
        _apply_header_style(ws)
        _apply_row_style(ws,min_row=2,max_row=ws.max_row,status_col_idx=status_col_idx)

        #sheet2需关注
        attention=df[df["status"].isin(["DIFF","UNMATCHED","NO_PMS_PRICE","PMS_ONLY"])]
        if not attention.empty:
            attention.to_excel(writer,sheet_name="需关注",index=False)
            ws2=writer.sheets["需关注"]
            _apply_header_style(ws2)
            _apply_row_style(ws2,min_row=2,max_row=ws2.max_row,status_col_idx=status_col_idx)


        #sheet3 摘要
        summary_rows=[
            ("总记录", stats["total"]),
            ("携程原始条数", total_orders),
            ("PMS原始条数", len(pms_records) if has_pms else 0),
            ("正常", stats["ok"]),
            ("佣金差异", stats["diff"]),
            ("未匹配", stats["unmatched"]),
            ("无PMS房价", stats["no_pms_price"]),
            ("PMS独有", stats["pms_only"]),
            ("佣金率", 0.15),
            ("容差(元)", 0.05),
        ]

        pd.DataFrame(summary_rows,columns=["指标","数值"]).to_excel(writer,sheet_name="摘要",index=False)

        ws3=writer.sheets["摘要"]
        _apply_header_style(ws3)

    # 删除上传的文件
    try:
        os.remove(ctrip_path)
        logger.info(f"已删除携程上传文件: {ctrip_path}")
    except OSError as e:
        logger.warning(f"[warn] 删除携程文件失败:{e}")

    if has_pms:
        try:
            os.remove(pms_path)
            logger.info(f"已删除PMS上传文件: {pms_path}")
        except OSError as e:
            logger.warning(f"[warn] 删除PMS文件失败:{e}")
    # 返回
    base_msg=(
        f"done: {out_path} "
        f"orders={total_orders} nights={int(total_nights)} comm={total_comm:,.2f}"
    )
    if has_pms:
        return (
                f"{base_msg} "
                f"ok={stats['ok']} diffs={stats['diff']} "
                f"unmatched={stats['unmatched']} pms_only={stats['pms_only']}"
        )

    return f"{base_msg} pms=skipped (parsed only)"