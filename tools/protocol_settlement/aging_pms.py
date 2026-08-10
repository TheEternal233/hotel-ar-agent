"""
M03: PMS应收账龄分析与坏账计提工具（协议客户对账）
"""
import os
import copy
from datetime import datetime
from pathlib import Path
from typing import  Dict, Any

try:
    from langchain.tools import tool
except ImportError:
    tool = None
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from . import generate_payment_notices
from .doc_parser_pms import read_pms_receivable, _parse_date


# ========== 路径配置 ==========
# 模板文件与工具脚本放在同级目录下
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "附件二 应收账龄分析表.xlsx")
# 输出目录：根目录下的 output 文件夹
# 当前脚本绝对路径
CURR_FILE = os.path.abspath(__file__)
# 当前脚本所在文件夹
CURR_DIR = os.path.dirname(CURR_FILE)
# 向上回溯2层，抵达项目根目录（根据你的层级修改数字！）
PROJECT_ROOT = os.path.dirname(os.path.dirname(CURR_DIR))
# 项目根目录/output
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")

# ========== 样式常量（与原有aging.py保持一致）==========
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

# ========== 账龄分段定义 ==========
AGING_BRACKETS = [
    (range(1, 31),   "1-30天",   "1-30"),
    (range(31, 61),  "31-60天",  "31-60"),
    (range(61, 91),  "61-90天",  "61-90"),
    (range(91, 121), "91-120天", "91-120"),
    (range(121, 151), "121-150天", "121-150"),
    (range(151, 181), "151-180天", "151-180"),
    (range(181, 99999), "180天以上", "180+"),
]

TEMPLATE_COLS = ["账号", "账户名称", "1-30天", "31-60天", "61-90天",
                 "91-120天", "121-150天", "151-180天", "180天以上", "合计金额"]


def _get_bracket_key(days: int) -> str:
    """根据逾期天数返回账龄段内部key"""
    if days < 1:
        return ""
    for brange, _, bkey in AGING_BRACKETS:
        if days in brange:
            return bkey
    return "180+"

def _analyze_pms_data(source_path: str, as_of_date: datetime) -> Dict[str, Any]:
    """对PMS应收账务列表进行账龄分析（内部核心逻辑）"""
    raw_records = read_pms_receivable(source_path)

    valid_records = []
    for rec in raw_records:
        # 跳过无协议单位的记录
        if not rec.get("effective_corp"):
            continue
        # 跳过汇总行（类型为空），避免与子明细重复
        if not rec.get("type"):
            continue
        # 只取借方记录（应收增加）
        if rec.get("type") != "借方":
            continue
        # 必须有有效日期
        if rec.get("date") is None:
            continue

        amt = rec.get("balance", 0)
        # 只保留正数金额
        if amt <= 0:
            continue

        valid_records.append({
            "bill_no": rec.get("bill_no", "").strip(),
            "corp": rec["effective_corp"],
            "date": rec["date"],
            "amount": amt,
        })

    # 按协议单位分组并计算账龄
    customers = {}
    for rec in valid_records:
        corp = rec["corp"]
        days = (as_of_date - rec["date"]).days

        if corp not in customers:
            customers[corp] = {
                "account_no": "",
                "bill_nos": [],
                "amounts": {k: 0.0 for _, _, k in AGING_BRACKETS},
                "total": 0.0,
            }

        bkey = _get_bracket_key(days)
        if not bkey:
            continue

        customers[corp]["amounts"][bkey] += rec["amount"]
        customers[corp]["total"] += rec["amount"]
        customers[corp]["bill_nos"].append(rec["bill_no"])

        if not customers[corp]["account_no"] and rec["bill_no"]:
            customers[corp]["account_no"] = rec["bill_no"]

    # 计算总计
    grand_total = {k: 0.0 for _, _, k in AGING_BRACKETS}
    grand_total["total"] = 0.0
    for corp, data in customers.items():
        for k in grand_total:
            if k != "total":
                grand_total[k] += data["amounts"].get(k, 0)
        grand_total["total"] += data["total"]

    return {
        "as_of_date": as_of_date,
        "customer_count": len(customers),
        "total_amount": grand_total["total"],
        "customers": customers,
        "grand_total": grand_total,
    }


def _write_aging_report(result: Dict[str, Any], template_path: str, output_path: str) -> str:
    """将账龄分析结果写入模板文件"""
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    wb = load_workbook(template_path)
    try:
        ws = wb.active

        # 定位Total行
        total_row = None
        for row_idx in range(2, ws.max_row + 1):
            cell_b = ws.cell(row=row_idx, column=2)
            if cell_b.value and str(cell_b.value).strip().lower() in ["total", "合计"]:
                total_row = row_idx
                break

        if total_row is None:
            total_row = ws.max_row + 1

        # 清空旧数据
        for row_idx in range(2, total_row):
            for col_idx in range(1, 11):
                ws.cell(row=row_idx, column=col_idx).value = None

        # 写入数据
        customers = result["customers"]
        sorted_customers = sorted(customers.items(), key=lambda x: -x[1]["total"])
        data_start_row = 2

        for idx, (corp_name, data) in enumerate(sorted_customers):
            row_idx = data_start_row + idx

            ws.cell(row=row_idx, column=1, value=data["account_no"])
            ws.cell(row=row_idx, column=2, value=corp_name)

            for col_offset, (_, bname, bkey) in enumerate(AGING_BRACKETS, start=3):
                val = data["amounts"].get(bkey, 0)
                cell = ws.cell(row=row_idx, column=col_offset, value=round(val, 2) if val else 0)
                cell.border = THIN_BORDER
                cell.number_format = '#,##0.00'
                if bkey in ["91-120", "121-150", "151-180"] and val > 0:
                    cell.fill = YELLOW_FILL
                elif bkey == "180+" and val > 0:
                    cell.fill = RED_FILL

            total_cell = ws.cell(row=row_idx, column=10, value=round(data["total"], 2))
            total_cell.border = THIN_BORDER
            total_cell.number_format = '#,##0.00'
            total_cell.font = Font(bold=True)

            for c in [1, 2]:
                ws.cell(row=row_idx, column=c).border = THIN_BORDER

        # 更新Total行
        new_total_row = data_start_row + len(sorted_customers)

        if total_row != new_total_row and total_row is not None:
            for col_idx in range(1, 11):
                old_cell = ws.cell(row=total_row, column=col_idx)
                new_cell = ws.cell(row=new_total_row, column=col_idx)
                new_cell.value = old_cell.value
                if old_cell.font:
                    new_cell.font = Font(
                        name=old_cell.font.name,
                        size=old_cell.font.size,
                        bold=old_cell.font.bold,
                        italic=old_cell.font.italic,
                        color=old_cell.font.color,
                    )
                if old_cell.fill and old_cell.fill.fill_type:
                    new_cell.fill = PatternFill(
                        start_color=old_cell.fill.start_color.rgb if old_cell.fill.start_color else None,
                        end_color=old_cell.fill.end_color.rgb if old_cell.fill.end_color else None,
                        fill_type=old_cell.fill.fill_type,
                    )
                if old_cell.border:
                    new_cell.border = copy.copy(old_cell.border)
                old_cell.value = None

        total_row = new_total_row
        ws.cell(row=total_row, column=1, value="").border = THIN_BORDER
        ws.cell(row=total_row, column=2, value="Total").font = Font(bold=True)
        ws.cell(row=total_row, column=2).border = THIN_BORDER
        ws.cell(row=total_row, column=2).fill = TOTAL_FILL

        for col_idx in range(3, 11):
            cell = ws.cell(row=total_row, column=col_idx)
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row - 1})"
            cell.font = Font(bold=True)
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
            cell.number_format = '#,##0.00'

        # 调整列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 35
        for c in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[c].width = 14

        wb.save(output_path)
        return output_path
    finally:
        wb.close()


# ========== 前端可调用的Tool接口 ==========

if tool is not None:
    @tool
    def aging_analysis(receivable_path: str, as_of_date: str = "",keep_source:bool=False,generate_notice: bool = True) -> str:
        """PMS应收账龄分析：读取PMS应收账务列表xlsx，按协议单位分组汇总，以交易日期计算账龄，
        使用余额作为未核销金额，生成账龄分析报表。
        默认同时生成付款通知书
        """
        if not os.path.exists(receivable_path):
            return f"错误：源文件不存在 {receivable_path}"
        if not os.path.exists(TEMPLATE_PATH):
            return f"错误：模板文件不存在 {TEMPLATE_PATH}"

        # 解析截止日期
        if not as_of_date:
            as_of_date_dt = datetime.now()
        else:
            as_of_date_dt = _parse_date(as_of_date) or datetime.now()

        # 执行分析
        result = _analyze_pms_data(receivable_path, as_of_date_dt)

        # 确保输出目录存在
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        # 生成输出路径
        output_path = os.path.join(OUTPUT_DIR, f"应收账龄分析报表_{as_of_date_dt.strftime('%Y%m%d')}.xlsx")

        # 写入报表
        _write_aging_report(result, TEMPLATE_PATH, output_path)

        # 默认生成付款通知书
        notice_result=""
        if generate_notice and os.path.exists(receivable_path):
            try:
                # 推断账期月份
                row_records=read_pms_receivable(receivable_path)
                dates=[r.get("date") for r in row_records if r.get("date") and r.get("type")=="借方"]
                if dates:
                    notice_month=max(dates).strftime("%Y-%m")
                else:
                    notice_month=datetime.now().strftime("%Y-%m")
                notice_result=generate_payment_notices(
                    receivable_path=receivable_path,
                    notice_month=notice_month,
                )
            except Exception as e:
                notice_result=f"付款通知书生成失败:{e}"

        if not keep_source:
            # 处理完之后删除上传的源文件
            try:
                p=Path(receivable_path).resolve()
                base=Path(PROJECT_ROOT).resolve()
                if str(p).startswith(str(base)) and p.exists():
                    os.remove(receivable_path)
            except Exception:
                pass

        # 构建返回摘要
        lines = [
            f"账龄分析完成: {output_path}",
            f"截止日期: {as_of_date_dt.strftime('%Y-%m-%d')}",
            f"客户数: {result['customer_count']}",
            f"应收总额: {result['total_amount']:,.2f}",
            "",
            "账龄分布:",
        ]

        for _, bname, bkey in AGING_BRACKETS:
            val = result["grand_total"].get(bkey, 0)
            lines.append(f"  {bname}: {val:,.2f}")

        lines.append("")
        lines.append("客户明细:")
        for corp_name, data in sorted(result["customers"].items(), key=lambda x: -x[1]["total"]):
            lines.append(f"  {corp_name:30} 合计: {data['total']:>12,.2f}  账号: {data['account_no']}")

        if notice_result:
            lines.extend([
                "",
                "=" * 60,
                "                   付款通知书生成结果",
                "=" * 60,
                notice_result,
            ])
        return "\n".join(lines)

    @tool
    def aging_and_notice(
        receivable_path: str,
        as_of_date: str = "",
        notice_month: str = "",
        output_dir: str = "",
        notice_date: str = "",
        due_date: str = "",
    ) -> str:
        """PMS账龄分析与付款通知书联合生成
        先执行应收账龄分析，再基于同一数据源为各协议客户生成付款通知书。
        全部完成后统一删除源文件。
        """
        # 1. 账龄分析（保留源文件）
        aging_result = aging_analysis.invoke({
            "receivable_path": receivable_path,
            "as_of_date": as_of_date,
            "keep_source": True,
            "generate_notice": False, #避免重复生成
        })
        if aging_result.startswith("错误"):
            return aging_result
        # 2. 推断付款通知书账期
        if not notice_month:
            raw_records = read_pms_receivable(receivable_path)
            dates = [r.get("date") for r in raw_records if r.get("date") and r.get("type") == "借方"]
            if dates:
                notice_month = max(dates).strftime("%Y-%m")
            else:
                notice_month = datetime.now().strftime("%Y-%m")
        # 3. 生成付款通知书
        notice_result = generate_payment_notices(
            receivable_path=receivable_path,
            notice_month=notice_month,
            output_dir=output_dir,
            notice_date=notice_date or None,
            due_date=due_date or None,
        )
        # 4. 统一清理源文件
        try:
            if os.path.exists(receivable_path):
                os.remove(receivable_path)
        except Exception:
            pass

        # 5. 组合返回
        return "\n".join([
            "=" * 70,
            "                     账龄分析 + 付款通知书 联合生成报告",
            "=" * 70, "",
            aging_result, "",
            "-" * 70, "",
            notice_result, "",
            "=" * 70,
            "全部任务完成，源文件已清理。",
            "=" * 70,
        ])

else:
    aging_analysis = None
    aging_and_notice = None