"""付款通知书 — 通用工具函数"""

import copy
from datetime import datetime, timedelta
from typing import Optional, Tuple

from openpyxl.styles import Font, Alignment, PatternFill, Border
from openpyxl.utils import get_column_letter

from .config import TemplateRows, DetailCols, DEFAULT_ADJUSTMENT


def resolve_notice_month(notice_month: str) -> Tuple[datetime, datetime, str]:
    """解析账期字符串，返回(月初, 月末, 显示文本)"""
    cleaned = notice_month.strip().replace("年", "-").replace("月", "")
    parts = cleaned.split("-")
    if len(parts) != 2:
        raise ValueError(f"账期格式错误: {notice_month}")

    year, month = int(parts[0]), int(parts[1])
    month_start = datetime(year, month, 1)
    next_month = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
    month_end = next_month - timedelta(seconds=1)
    return month_start, month_end, f"{year}年{month}月消费"


def is_in_month(dt: Optional[datetime], month_start: datetime, month_end: datetime) -> bool:
    return dt is not None and month_start <= dt <= month_end


def calc_open_balance(records, corp: str, month_start: datetime) -> float:
    """计算上期余额（截至上月末的未核销借方余额）"""
    total = 0.0
    for rec in records:
        if rec.get("effective_corp") != corp or rec.get("type") != "借方":
            continue
        rec_date = rec.get("date")
        if rec_date is None or rec_date >= month_start:
            continue
        bal = rec.get("balance", 0)
        if bal > 0:
            total += bal
    return round(total, 2)


def copy_cell_style(src, dst):
    """复制单元格样式"""
    if src.font:
        dst.font = Font(
            name=src.font.name, size=src.font.size, bold=src.font.bold,
            italic=src.font.italic, underline=src.font.underline,
            strike=src.font.strike, color=src.font.color,
        )
    if src.fill and src.fill.fill_type:
        dst.fill = PatternFill(
            start_color=src.fill.start_color.rgb if src.fill.start_color else None,
            end_color=src.fill.end_color.rgb if src.fill.end_color else None,
            fill_type=src.fill.fill_type,
        )
    if src.border:
        dst.border = Border(
            left=copy.copy(src.border.left), right=copy.copy(src.border.right),
            top=copy.copy(src.border.top), bottom=copy.copy(src.border.bottom),
        )
    if src.alignment:
        dst.alignment = Alignment(
            horizontal=src.alignment.horizontal, vertical=src.alignment.vertical,
            wrap_text=src.alignment.wrap_text,
        )
    dst.number_format = src.number_format


def insert_rows_with_style(ws, insert_after_row: int, count: int):
    """插入新行并复制样式，正确处理合并单元格"""
    merges_to_shift = [mr for mr in list(ws.merged_cells.ranges) if mr.min_row > insert_after_row]
    template_merges = [mr for mr in list(ws.merged_cells.ranges)
                       if mr.min_row <= insert_after_row <= mr.max_row]

    for mr in merges_to_shift:
        ws.unmerge_cells(str(mr))

    insert_at = insert_after_row + 1
    ws.insert_rows(insert_at, count)

    for offset in range(count):
        new_row = insert_at + offset
        for col_idx in range(1, ws.max_column + 1):
            src = ws.cell(row=insert_after_row, column=col_idx)
            dst = ws.cell(row=new_row, column=col_idx)
            copy_cell_style(src, dst)
            dst.value = None

    for offset in range(count):
        new_row = insert_at + offset
        for mr in template_merges:
            rs = new_row - insert_after_row
            rng = (f"{get_column_letter(mr.min_col)}{mr.min_row + rs}:"
                   f"{get_column_letter(mr.max_col)}{mr.max_row + rs}")
            try:
                ws.merge_cells(rng)
            except ValueError:
                pass

    for mr in merges_to_shift:
        rng = (f"{get_column_letter(mr.min_col)}{mr.min_row + count}:"
               f"{get_column_letter(mr.max_col)}{mr.max_row + count}")
        try:
            ws.merge_cells(rng)
        except ValueError:
            pass


def update_sum_formula(ws, total_row: int, first_data_row: int, last_data_row: int):
    """更新合计行SUM公式"""
    ws.cell(row=total_row, column=DetailCols.AMOUNT).value = (
        f"=SUM(G{first_data_row}:G{last_data_row})"
    )