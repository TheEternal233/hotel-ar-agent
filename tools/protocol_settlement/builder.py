"""付款通知书 — 数据构建与模板填充"""

import os
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, List, Optional
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook

from .config import (
    TemplateRows, DetailCols, DEFAULT_ADJUSTMENT, DEFAULT_DUE_DAYS,
    NOTICE_TEMPLATE_PATH, THIN_BORDER,
)
from .utils import (
    is_in_month, calc_open_balance,
    insert_rows_with_style, update_sum_formula,
)


def filter_records_by_month(records, month_start: datetime, month_end: datetime):
    """筛选指定月份的借方记录"""
    filtered = []
    for rec in records:
        if not rec.get("effective_corp") or rec.get("type") != "借方":
            continue
        if not is_in_month(rec.get("date"), month_start, month_end):
            continue
        amt = rec.get("debit", 0) or rec.get("amount", 0)
        if amt > 0:
            filtered.append(rec)
    return filtered


def build_corp_summary(records, month_start: datetime, month_end: datetime,
                       open_balances: Optional[Dict[str, float]] = None):
    """按协议单位分组，构建付款通知书数据"""
    month_records = filter_records_by_month(records, month_start, month_end)
    corp_groups: Dict[str, List[Dict]] = {}
    for rec in month_records:
        corp = rec["effective_corp"]
        corp_groups.setdefault(corp, []).append(rec)

    summary = {}
    for corp, corp_records in corp_groups.items():
        open_balance = float(open_balances[corp]) if (open_balances and corp in open_balances) else \
                       calc_open_balance(records, corp, month_start)

        details = []
        for rec in sorted(corp_records, key=lambda r: r.get("date") or datetime.min):
            details.append({
                "date": rec.get("date"),
                "conf_no": rec.get("ext_order") or rec.get("order_no") or rec.get("bill_no", ""),
                "guest_name": rec.get("name_desc", ""),
                "amount": round(rec.get("debit", 0) or rec.get("amount", 0), 2),
            })

        detail_total = round(sum(d["amount"] for d in details), 2)
        adjustment = DEFAULT_ADJUSTMENT
        grand_total = round(open_balance + detail_total + adjustment, 2)

        summary[corp] = {
            "open_balance": open_balance,
            "details": details,
            "detail_total": detail_total,
            "adjustment": adjustment,
            "grand_total": grand_total,
        }
    return summary


def fill_notice_template(corp_name: str, data: Dict[str, Any],
                         template_path: str, output_path: str,
                         notice_month_display: str, notice_date: datetime,
                         due_date: Optional[datetime] = None,
                         adjustment: Optional[float] = None) -> str:
    """填充单个协议单位的付款通知书模板"""
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    try:
        ws = wb.active
        tr = TemplateRows

        # 抬头
        ws.cell(row=tr.TO_CLIENT, column=1, value=f"致：{corp_name}")
        ws.cell(row=tr.DATE, column=7, value=notice_date)
        ws.cell(row=tr.DATE, column=7).number_format = "YYYY-MM-DD"
        ws.cell(row=tr.SUBJECT, column=2, value=notice_month_display)

        # 上期余额
        ws.cell(row=tr.OPEN_BALANCE, column=DetailCols.GUEST_NAME, value="上期余额")
        ws.cell(row=tr.OPEN_BALANCE, column=DetailCols.AMOUNT, value=data["open_balance"])
        ws.cell(row=tr.OPEN_BALANCE, column=DetailCols.AMOUNT).number_format = "#,##0.00"

        # 消费明细
        details = data["details"]
        capacity = tr.DETAIL_END - tr.DETAIL_START + 1
        extra_rows = 0
        if len(details) > capacity:
            extra_rows = len(details) - capacity
            insert_rows_with_style(ws, tr.DETAIL_END, extra_rows)

        adj_row = tr.ADJUSTMENT + extra_rows
        total_row = tr.TOTAL + extra_rows
        due_row = tr.PAYMENT_DUE + extra_rows

        for idx, detail in enumerate(details):
            row = tr.DETAIL_START + idx
            ws.cell(row=row, column=DetailCols.DATE, value=detail["date"]).number_format = "YYYY-MM-DD"
            ws.cell(row=row, column=DetailCols.CONF_NO, value=detail["conf_no"])
            ws.cell(row=row, column=DetailCols.GUEST_NAME, value=detail["guest_name"])
            ws.cell(row=row, column=DetailCols.AMOUNT, value=detail["amount"]).number_format = "#,##0.00"

        if extra_rows == 0 and len(details) < capacity:
            for row in range(tr.DETAIL_START + len(details), tr.DETAIL_END + 1):
                for col in [DetailCols.DATE, DetailCols.CONF_NO, DetailCols.GUEST_NAME, DetailCols.AMOUNT]:
                    ws.cell(row=row, column=col).value = None

        # 小数调整
        adj = adjustment if adjustment is not None else data["adjustment"]
        ws.cell(row=adj_row, column=DetailCols.GUEST_NAME, value="       小 数 调 整")
        ws.cell(row=adj_row, column=DetailCols.AMOUNT, value=adj).number_format = "#,##0.00"

        # 合计
        update_sum_formula(ws, total_row, tr.OPEN_BALANCE, adj_row)
        ws.cell(row=total_row, column=6, value="CNY")
        ws.cell(row=total_row, column=DetailCols.AMOUNT).number_format = "#,##0.00"

        # 付款期限
        if due_date is None:
            due_date = notice_date + timedelta(days=DEFAULT_DUE_DAYS)
        ws.cell(row=due_row, column=1, value=f"Payment Due Date: {due_date.strftime('%Y-%m-%d')}")

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        logo_path = Path(__file__).parent / "logo.png"
        if logo_path.exists():
            try:
                logo = XLImage(str(logo_path))
                logo.width = 180
                logo.height = 60
                ws.add_image(logo, "C3")
            except ImportError:
                pass
        wb.save(output_path)
        return output_path
    finally:
        wb.close()