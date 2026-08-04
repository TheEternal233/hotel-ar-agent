"""PMS应收账务解析模块 —— 基于doc_parser标准接口扩展

专门处理清远酒店PMS系统导出的「应收账务列表」Excel，提供：
- 列映射与读取
- 协议单位提取（支持从转账注释TAr字段提取）
- 日期与金额解析
"""
import os
import re
from datetime import datetime
from typing import Optional, List, Dict, Any

import openpyxl

# ========== PMS应收账务列映射 ==========
PMS_RECEIVABLE_MAPPING = {
    "bill_no":       ["账单号", "bill_no", "账单编号"],
    "type":          ["类型", "type", "借贷类型"],
    "date":          ["日期", "date", "业务日期", "交易日期"],
    "checkout_time": ["结账时间", "checkout_time"],
    "name_desc":     ["姓名/描述", "姓名", "描述", "name_desc"],
    "room":          ["房号", "room", "room_no"],
    "amount":        ["金额", "amount", "交易金额"],
    "debit":         ["借方", "debit", "借方金额"],
    "credit":        ["贷方", "credit", "贷方金额"],
    "written_off":   ["已核销", "written_off", "核销"],
    "balance":       ["余额", "balance", "当前余额"],
    "finance_note":  ["财务备注", "finance_note"],
    "note":          ["备注", "note"],
    "transfer_note": ["转账注释", "transfer_note", "转账备注"],
    "checkout_bill": ["结账单号", "checkout_bill"],
    "dispute":       ["争议余额", "dispute"],
    "order_no":      ["订单号", "order_no", "订单编号"],
    "ext_order":     ["外部订单号", "ext_order"],
    "central_order": ["中央订单号", "central_order"],
    "corp":          ["协议单位", "corp", "协议公司", "客户"],
    "order_remark":  ["订单备注", "order_remark"],
    "operator":      ["入账操作员", "operator", "操作员"],
}



# ========== 内部工具函数 ==========

def _open_workbook(path:str,sheet_name:str=None):
    """打开Excel工作簿，返回(workbook,worksheet)"""
    if not os.path.exists(path):
        raise FileNotFoundError(f"文件不存在:{path}")
    wb=openpyxl.load_workbook(path,data_only=True)
    ws=wb[sheet_name] if sheet_name else wb.active
    return wb,ws

def _get_headers(ws,header_row:int=1)->list[str]:
    """读取指定行的表头"""
    row_vals=[c.value for c in next(ws.iter_rows(min_row=header_row,max_row=header_row))]
    return [str(v).strip() if v is not None else "" for v in row_vals]

def _parse_date(val)->Optional[datetime]:
    """解析日期，支持多种格式"""
    if isinstance(val,datetime):
        return val
    if isinstance(val,str) and val.strip():
        for fmt in ["%Y-%m-%d","%Y-%m-%d %H:%M:%S","%Y%m%d","%Y/%m/%d"]:
            try:
                return datetime.strptime(val.strip(),fmt)
            except ValueError:
                pass
    return None

def _parse_amount(val)->float:
    """解析金额，空值返回0"""
    if val is None:
        return 0.0
    if isinstance(val,(int,float)):
        return float(val)
    if isinstance(val,str):
        try:
            return float(val.replace(",","").replace("¥","").strip())
        except ValueError:
            return 0.0
    return 0.0

def extract_corp_from_note(note:str)->Optional[str]:
    """
    从转账注释中提取协议单位（TAr字段）

    示例: "FGst:113706黄笠舟;TAr:广州长颈鹿旅行社|129174..."
    返回: "广州长颈鹿旅行社"
    """
    if not note:
        return None
    # 匹配TAr:XXX 或TAr:xxx/
    m=re.search(r'TAr:([^;|]+)', note)
    if m:
        return m.group(1).strip()

    return None


def normalize_corp_name(name:str)->str:
    """
    读取PMS应收账务列表Excel，返回标准化记录列表

    每条记录包含以下字段：
        bill_no, type, date, checkout_time, name_desc, room,
        amount, debit, credit, written_off, balance,
        finance_note, note, transfer_note, checkout_bill,
        dispute, order_no, ext_order, central_order,
        corp, order_remark, operator,
        corp_extracted（从转账注释提取的协议单位）
    """

    if not name:
        return ""
    name=name.strip()
    #去除尾部"付款"字样
    name=re.sub(r'\s*付款\s*$', '', name)
    return name.strip()


# ========== 核心读取函数 ==========

def read_pms_receivable(path: str, header_row: int = 1, sheet_name: str = None) -> List[Dict[str, Any]]:
    """读取PMS应收账务列表Excel，返回标准化记录列表

    每条记录包含以下字段：
        bill_no, type, date, checkout_time, name_desc, room,
        amount, debit, credit, written_off, balance,
        finance_note, note, transfer_note, checkout_bill,
        dispute, order_no, ext_order, central_order,
        corp, order_remark, operator,
        corp_extracted（从转账注释提取的协议单位）
    """
    wb, ws = _open_workbook(path, sheet_name)
    headers = _get_headers(ws, header_row)
    headers_lower = [h.lower() for h in headers]

    # 建立字段到列索引的映射
    field_to_col = {}
    for std_name, keywords in PMS_RECEIVABLE_MAPPING.items():
        for kw in keywords:
            kw_lower = kw.lower()
            for idx, hl in enumerate(headers_lower):
                if kw_lower in hl:
                    field_to_col[std_name] = idx
                    break
            if std_name in field_to_col:
                break

    records = []
    start_row = header_row + 1
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        # 跳过全空行
        if all(v is None for v in row):
            continue

        rec = {}
        for std_name, col_idx in field_to_col.items():
            val = row[col_idx] if col_idx < len(row) else None
            rec[std_name] = val

        # 类型转换
        rec["date"] = _parse_date(rec.get("date"))
        rec["debit"] = _parse_amount(rec.get("debit"))
        rec["credit"] = _parse_amount(rec.get("credit"))
        rec["written_off"] = _parse_amount(rec.get("written_off"))
        rec["balance"] = _parse_amount(rec.get("balance"))
        rec["amount"] = _parse_amount(rec.get("amount"))
        rec["dispute"] = _parse_amount(rec.get("dispute"))

        # 字符串清理
        for k in ["bill_no", "type", "name_desc", "room", "finance_note",
                  "note", "transfer_note", "checkout_bill", "order_no",
                  "ext_order", "central_order", "corp", "order_remark", "operator"]:
            if rec.get(k) is not None:
                rec[k] = str(rec[k]).strip()
            else:
                rec[k] = ""

        # 从转账注释提取协议单位
        rec["corp_extracted"] = extract_corp_from_note(rec.get("transfer_note", ""))

        # 确定有效协议单位（优先使用corp列，其次从转账注释提取）
        effective_corp = rec["corp"] if rec["corp"] else rec["corp_extracted"]
        rec["effective_corp"] = normalize_corp_name(effective_corp) if effective_corp else ""

        records.append(rec)

    wb.close()
    return records



def get_pms_info(path:str, sheet_name: str = None) -> Dict[str, Any]:
    """获取PMS应收财务文件的基本信息"""
    wb,ws=_open_workbook(path, sheet_name)
    headers = _get_headers(ws,1)
    rows=sum(1 for _ in ws.iter_rows(min_row=2))
    wb.close()
    return {
        "path": path,
        "filename": os.path.basename(path),
        "sheet": ws.title,
        "cols":len(headers),
        "rows":rows,
        "headers": headers,
    }


def validate_pms_receivable(path:str, sheet_name: str = None) -> tuple:
    """验证PMS应收帐务文件是否包含必要的列"""
    required=["账单号", "类型", "日期", "借方", "贷方", "余额", "协议单位"]
    try:
        wb,ws=_open_workbook(path, sheet_name)
        headers = _get_headers(ws,1)
        headers_str=[str(h) for h in headers]
        missing=[c for c in required if not any(c in h for h in headers_str)]
        wb.close()
        if missing:
            return False, f"缺少必要列: {missing} (现有: {headers_str})"
        return True,f"验证通过: {len(headers_str)} 列"
    except Exception as e:
        return False,f"验证失败:{e}"

































































