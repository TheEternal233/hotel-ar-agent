"""M04: PMS付款通知书生成工具（协议客户对账）

基于PMS「应收账务列表」数据源，为每个协议客户生成独立的付款通知书：
1. 读取PMS应收账务明细
2. 按协议单位分组，筛选指定账期内的消费记录
3. 计算上期余额（自动或手动）
4. 填充付款通知书模板（附件三），含客户名称、账期、消费明细、金额汇总
5. 为每个协议单位输出独立的Excel文件
"""

import os
import copy
import re
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .doc_parser_pms import read_pms_receivable, _parse_date


# ============================================================================
# 配置常量
# ============================================================================

# 模板文件路径（与脚本同级目录）
NOTICE_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "附件三 OTA 月度对账底稿付款通知书 模板.xlsx"
)

# 输出目录（沿用aging_pms.py的PROJECT_ROOT逻辑）
_CURR_FILE = os.path.abspath(__file__)
_CURR_DIR = os.path.dirname(_CURR_FILE)
_PROJECT_ROOT = os.path.dirname(os.path.dirname(_CURR_DIR))
DEFAULT_OUTPUT_DIR = os.path.join(_PROJECT_ROOT, "output", "付款通知书")

# 模板中的固定行号映射（基于附件三模板结构）
class _TemplateRows:
    """模板行号常量，便于维护"""
    TO_CLIENT = 10          # A10: 致：XXX
    DATE = 11               # G11: 日期
    SUBJECT = 14            # B14: 202X年X月消费
    HEADER = 17             # 表头行
    OPEN_BALANCE = 18       # 上期余额行（C18=上期余额, G18=金额）
    DETAIL_START = 19       # 明细起始行
    DETAIL_END = 33         # 明细结束行（模板默认15行容量）
    ADJUSTMENT = 34         # 小数调整行（C34=小数调整, G34=-0.01）
    TOTAL = 36              # 合计行（F36=CNY, G36=SUM(G18:G34)）
    PAYMENT_DUE = 49        # 付款期限行

# 明细区域列映射
class _DetailCols:
    """明细区域列号常量"""
    DATE = 1        # A列: 日期
    CONF_NO = 2     # B列: Conf NO. 预订号
    GUEST_NAME = 3  # C列: GUEST NAME 客人姓名（C-D合并）
    AMOUNT = 7      # G列: 金额

# 默认小数调整金额（与模板保持一致）
DEFAULT_ADJUSTMENT = -0.01

# 默认付款期限天数（自通知书日期起算）
DEFAULT_DUE_DAYS = 30

# 样式常量
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


# ============================================================================
# 内部工具函数
# ============================================================================

def _resolve_notice_month(notice_month: str) -> Tuple[datetime, datetime, str]:
    """解析账期字符串，返回(月初, 月末, 显示文本)

    """
    # 统一格式：支持 "2026-07" 或 "2026年7月"
    cleaned = notice_month.strip().replace("年", "-").replace("月", "")
    parts = cleaned.split("-")
    if len(parts) != 2:
        raise ValueError(f"账期格式错误: {notice_month}，应为 YYYY-MM 或 YYYY年MM月")

    year = int(parts[0])
    month = int(parts[1])

    month_start = datetime(year, month, 1)
    # 计算下月1日再减1秒
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    month_end = next_month - timedelta(seconds=1)

    display_text = f"{year}年{month}月消费"
    return month_start, month_end, display_text


def _is_in_month(dt: Optional[datetime], month_start: datetime, month_end: datetime) -> bool:
    """判断日期是否落在指定月份内"""
    if dt is None:
        return False
    return month_start <= dt <= month_end


def _filter_records_by_month(
    records: List[Dict[str, Any]],
    month_start: datetime,
    month_end: datetime
) -> List[Dict[str, Any]]:
    """从PMS记录中筛选指定月份的借方记录

    筛选条件：
    - 有有效协议单位（effective_corp）
    - 类型为"借方"（应收增加）
    - 交易日期在指定月份内
    - 金额大于0
    """
    filtered = []
    for rec in records:
        if not rec.get("effective_corp"):
            continue
        if rec.get("type") != "借方":
            continue
        rec_date = rec.get("date")
        if not _is_in_month(rec_date, month_start, month_end):
            continue
        amt = rec.get("debit", 0) or rec.get("amount", 0)
        if amt <= 0:
            continue

        filtered.append(rec)
    return filtered


def _calc_open_balance(
    records: List[Dict[str, Any]],
    corp: str,
    month_start: datetime
) -> float:
    """计算指定协议单位的上期余额（截至上月末的未核销借方余额）

    逻辑：取该协议单位在 notice_month 之前的所有借方记录，
          以 balance 字段作为未核销金额累加。

    """
    total = 0.0
    for rec in records:
        if rec.get("effective_corp") != corp:
            continue
        if rec.get("type") != "借方":
            continue
        rec_date = rec.get("date")
        if rec_date is None or rec_date >= month_start:
            continue
        # 使用余额作为未核销金额
        bal = rec.get("balance", 0)
        if bal > 0:
            total += bal
    return round(total, 2)


def _copy_cell_style(src, dst):
    """复制单元格样式（字体、填充、边框、对齐、数字格式）"""
    if src.font:
        dst.font = Font(
            name=src.font.name,
            size=src.font.size,
            bold=src.font.bold,
            italic=src.font.italic,
            underline=src.font.underline,
            strike=src.font.strike,
            color=src.font.color,
        )
    if src.fill and src.fill.fill_type:
        dst.fill = PatternFill(
            start_color=src.fill.start_color.rgb if src.fill.start_color else None,
            end_color=src.fill.end_color.rgb if src.fill.end_color else None,
            fill_type=src.fill.fill_type,
        )
    if src.border:
        dst.border = Border(
            left=copy.copy(src.border.left),
            right=copy.copy(src.border.right),
            top=copy.copy(src.border.top),
            bottom=copy.copy(src.border.bottom),
        )
    if src.alignment:
        dst.alignment = Alignment(
            horizontal=src.alignment.horizontal,
            vertical=src.alignment.vertical,
            wrap_text=src.alignment.wrap_text,
            shrink_to_fit=src.alignment.shrink_to_fit,
        )
    dst.number_format = src.number_format


def _insert_rows_with_style(ws, insert_after_row: int, count: int):
    """在指定行之后插入若干新行，并复制上一行的样式，正确处理合并单元格

    注意：openpyxl 的 insert_rows 不会自动调整合并单元格的行号，
    因此需要手动删除旧合并单元格并重新创建。

    """
    # 1. 在插入前，收集所有需要下移的合并单元格（行号 > insert_after_row）
    merges_to_shift = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row > insert_after_row:
            merges_to_shift.append(merged_range)

    # 2. 记录模板行的合并单元格（用于复制到新行）
    template_row = insert_after_row
    template_merges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= template_row <= merged_range.max_row:
            template_merges.append(merged_range)

    # 3. 删除需要下移的合并单元格（避免插入时冲突）
    for merged_range in merges_to_shift:
        ws.unmerge_cells(str(merged_range))

    # 4. 执行插入
    insert_at = insert_after_row + 1
    ws.insert_rows(insert_at, count)

    # 5. 复制样式到新插入的行
    for offset in range(count):
        new_row = insert_at + offset
        for col_idx in range(1, ws.max_column + 1):
            src = ws.cell(row=template_row, column=col_idx)
            dst = ws.cell(row=new_row, column=col_idx)
            _copy_cell_style(src, dst)
            dst.value = None  # 清空值

    # 6. 为模板行的合并单元格在新行中创建副本
    for offset in range(count):
        new_row = insert_at + offset
        for merged_range in template_merges:
            row_shift = new_row - template_row
            new_min_row = merged_range.min_row + row_shift
            new_max_row = merged_range.max_row + row_shift
            new_range = (
                f"{get_column_letter(merged_range.min_col)}{new_min_row}:"
                f"{get_column_letter(merged_range.max_col)}{new_max_row}"
            )
            try:
                ws.merge_cells(new_range)
            except ValueError:
                pass  # 可能已存在，忽略

    # 7. 重新创建下移后的合并单元格（行号 + count）
    for merged_range in merges_to_shift:
        new_min_row = merged_range.min_row + count
        new_max_row = merged_range.max_row + count
        new_range = (
            f"{get_column_letter(merged_range.min_col)}{new_min_row}:"
            f"{get_column_letter(merged_range.max_col)}{new_max_row}"
        )
        try:
            ws.merge_cells(new_range)
        except ValueError:
            pass


def _update_sum_formula(ws, total_row: int, first_data_row: int, last_data_row: int):
    """更新合计行的SUM公式范围

    模板原始公式: =SUM(G18:G34)
    更新后: =SUM(G{first_data_row}:G{last_data_row})
    """
    total_cell = ws.cell(row=total_row, column=_DetailCols.AMOUNT)
    total_cell.value = f"=SUM(G{first_data_row}:G{last_data_row})"


# ============================================================================
# 核心逻辑：构建协议单位汇总数据
# ============================================================================

def _build_corp_summary(
    records: List[Dict[str, Any]],
    month_start: datetime,
    month_end: datetime,
    open_balances: Optional[Dict[str, float]] = None,
) -> Dict[str, Dict[str, Any]]:
    """构建各协议单位的付款通知书数据   """
    # 1. 筛选本月记录
    month_records = _filter_records_by_month(records, month_start, month_end)

    # 2. 按协议单位分组
    corp_groups: Dict[str, List[Dict[str, Any]]] = {}
    for rec in month_records:
        corp = rec["effective_corp"]
        corp_groups.setdefault(corp, []).append(rec)

    # 3. 构建每个协议单位的汇总数据
    summary = {}
    for corp, corp_records in corp_groups.items():
        # 上期余额
        if open_balances and corp in open_balances:
            open_balance = float(open_balances[corp])
        else:
            open_balance = _calc_open_balance(records, corp, month_start)

        # 消费明细（按日期排序）
        details = []
        for rec in sorted(corp_records, key=lambda r: r.get("date") or datetime.min):
            details.append({
                "date": rec.get("date"),
                "conf_no": rec.get("ext_order") or rec.get("order_no") or rec.get("bill_no", ""),
                "guest_name": rec.get("name_desc", ""),
                "amount": round(rec.get("debit", 0) or rec.get("amount", 0), 2),
            })

        detail_total = round(sum(d["amount"] for d in details), 2)

        # 小数调整：默认-0.01
        adjustment = DEFAULT_ADJUSTMENT

        # 总计 = 上期余额 + 本月消费合计 + 小数调整
        grand_total = round(open_balance + detail_total + adjustment, 2)

        summary[corp] = {
            "open_balance": open_balance,
            "details": details,
            "detail_total": detail_total,
            "adjustment": adjustment,
            "grand_total": grand_total,
        }

    return summary


# ============================================================================
# 核心逻辑：填充单个付款通知书模板
# ============================================================================

def _fill_notice_template(
    corp_name: str,
    data: Dict[str, Any],
    template_path: str,
    output_path: str,
    notice_month_display: str,
    notice_date: datetime,
    due_date: Optional[datetime] = None,
    adjustment: Optional[float] = None,
) -> str:
    """填充单个协议单位的付款通知书模板"""
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active
    tr = _TemplateRows

    # ---------- 1. 填充抬头信息 ----------
    # 致：客户名称
    ws.cell(row=tr.TO_CLIENT, column=1, value=f"致：{corp_name}")

    # 日期
    ws.cell(row=tr.DATE, column=7, value=notice_date)
    ws.cell(row=tr.DATE, column=7).number_format = "YYYY-MM-DD"

    # 关于/Subject
    ws.cell(row=tr.SUBJECT, column=2, value=notice_month_display)

    # ---------- 2. 填充上期余额 ----------
    ws.cell(row=tr.OPEN_BALANCE, column=_DetailCols.GUEST_NAME, value="上期余额")
    ws.cell(row=tr.OPEN_BALANCE, column=_DetailCols.AMOUNT, value=data["open_balance"])
    ws.cell(row=tr.OPEN_BALANCE, column=_DetailCols.AMOUNT).number_format = "#,##0.00"

    # ---------- 3. 填充消费明细 ----------
    details = data["details"]
    template_capacity = tr.DETAIL_END - tr.DETAIL_START + 1  # 15行

    # 如果明细超过模板容量，需要插入行
    extra_rows = 0
    if len(details) > template_capacity:
        extra_rows = len(details) - template_capacity
        # 在第 DETAIL_END 行之后插入（即在第33行后插入，第34行前）
        _insert_rows_with_style(ws, tr.DETAIL_END, extra_rows)

    # 计算实际行号（考虑插入行后的偏移）
    actual_adjustment_row = tr.ADJUSTMENT + extra_rows
    actual_total_row = tr.TOTAL + extra_rows
    actual_payment_due_row = tr.PAYMENT_DUE + extra_rows
    last_detail_row = tr.DETAIL_START + len(details) - 1

    # 填充明细数据
    for idx, detail in enumerate(details):
        row = tr.DETAIL_START + idx
        # 日期
        date_cell = ws.cell(row=row, column=_DetailCols.DATE, value=detail["date"])
        date_cell.number_format = "YYYY-MM-DD"
        # 预订号
        ws.cell(row=row, column=_DetailCols.CONF_NO, value=detail["conf_no"])
        # 客人姓名（C-D合并列，写入C列即可）
        ws.cell(row=row, column=_DetailCols.GUEST_NAME, value=detail["guest_name"])
        # 金额
        amt_cell = ws.cell(row=row, column=_DetailCols.AMOUNT, value=detail["amount"])
        amt_cell.number_format = "#,##0.00"

    # 清空未使用的明细行（仅当未插入行时）
    if extra_rows == 0 and len(details) < template_capacity:
        for row in range(tr.DETAIL_START + len(details), tr.DETAIL_END + 1):
            for col in [_DetailCols.DATE, _DetailCols.CONF_NO,
                        _DetailCols.GUEST_NAME, _DetailCols.AMOUNT]:
                ws.cell(row=row, column=col).value = None

    # ---------- 4. 填充小数调整 ----------
    adj = adjustment if adjustment is not None else data["adjustment"]
    ws.cell(row=actual_adjustment_row, column=_DetailCols.GUEST_NAME,
            value="       小 数 调 整")
    ws.cell(row=actual_adjustment_row, column=_DetailCols.AMOUNT, value=adj)
    ws.cell(row=actual_adjustment_row, column=_DetailCols.AMOUNT).number_format = "#,##0.00"

    # ---------- 5. 更新合计公式 ----------
    first_data_row = tr.OPEN_BALANCE  # 从第18行开始求和
    _update_sum_formula(ws, actual_total_row, first_data_row, actual_adjustment_row)
    ws.cell(row=actual_total_row, column=6, value="CNY")  # F列
    ws.cell(row=actual_total_row, column=_DetailCols.AMOUNT).number_format = "#,##0.00"

    # ---------- 6. 填充付款期限 ----------
    if due_date is None:
        due_date = notice_date + timedelta(days=DEFAULT_DUE_DAYS)
    ws.cell(row=actual_payment_due_row, column=1,
            value=f"Payment Due Date: {due_date.strftime('%Y-%m-%d')}")

    # ---------- 7. 保存 ----------
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()
    return output_path


# ============================================================================
# 前端接口
# ============================================================================

def generate_payment_notices(
    receivable_path: str,
    notice_month: str,
    output_dir: str = "",
    notice_date: Optional[str] = None,
    due_date: Optional[str] = None,
    open_balances: Optional[Dict[str, float]] = None,
    adjustment: Optional[float] = None,
    template_path: str = "",
) -> str:
    """生成付款通知书主入口

    读取PMS应收账务列表，按协议单位分组，为每个协议客户生成一份
    付款通知书Excel文件。

    """
    # ---------- 参数校验与解析 ----------
    if not os.path.exists(receivable_path):
        return f"错误：源文件不存在 {receivable_path}"

    tpl_path = template_path or NOTICE_TEMPLATE_PATH
    if not os.path.exists(tpl_path):
        return f"错误：模板文件不存在 {tpl_path}"

    out_dir = output_dir or DEFAULT_OUTPUT_DIR
    os.makedirs(out_dir, exist_ok=True)

    # 解析账期
    try:
        month_start, month_end, month_display = _resolve_notice_month(notice_month)
    except ValueError as e:
        return f"错误：{e}"

    # 解析日期
    notice_dt = _parse_date(notice_date) or datetime.now()
    due_dt = _parse_date(due_date) if due_date else None

    # ---------- 读取PMS数据 ----------
    raw_records = read_pms_receivable(receivable_path)
    if not raw_records:
        return "错误：PMS应收账务列表为空或读取失败"

    # ---------- 构建汇总数据 ----------
    summary = _build_corp_summary(
        raw_records, month_start, month_end, open_balances
    )

    if not summary:
        return f"未找到 {notice_month} 账期内的有效消费记录"

    # ---------- 逐客户生成通知书 ----------
    generated_files = []
    failed_corps = []

    for corp_name, data in sorted(summary.items()):
        # 文件名：付款通知书_客户名_账期.xlsx
        safe_name = re.sub(r'[\\/:*?"<>|]', "_", corp_name)
        filename = f"付款通知书_{safe_name}_{month_start.strftime('%Y%m')}.xlsx"
        output_path = os.path.join(out_dir, filename)

        try:
            _fill_notice_template(
                corp_name=corp_name,
                data=data,
                template_path=tpl_path,
                output_path=output_path,
                notice_month_display=month_display,
                notice_date=notice_dt,
                due_date=due_dt,
                adjustment=adjustment,
            )
            generated_files.append({
                "corp": corp_name,
                "path": output_path,
                "details_count": len(data["details"]),
                "open_balance": data["open_balance"],
                "grand_total": data["grand_total"],
            })
        except Exception as e:
            failed_corps.append(f"{corp_name}: {e}")

    # ---------- 构建返回摘要 ----------
    lines = [
        "付款通知书生成完成",
        f"账期: {notice_month} ({month_display})",
        f"通知书日期: {notice_dt.strftime('%Y-%m-%d')}",
        f"输出目录: {out_dir}",
        f"成功生成: {len(generated_files)} 份",
        "",
        "=" * 60,
        "生成文件清单:",
        "=" * 60,
    ]

    for item in generated_files:
        lines.append(f"\n  客户: {item['corp']}")
        lines.append(f"  文件: {os.path.basename(item['path'])}")
        lines.append(f"  上期余额: {item['open_balance']:,.2f}")
        lines.append(f"  本月明细: {item['details_count']} 笔")
        lines.append(f"  应付总额: {item['grand_total']:,.2f}")

    if failed_corps:
        lines.append("\n")
        lines.append("=" * 60)
        lines.append(f"失败 ({len(failed_corps)} 个):")
        lines.append("=" * 60)
        for fail in failed_corps:
            lines.append(f"  {fail}")

    return "\n".join(lines)




try:
    from langchain.tools import tool

    @tool
    def payment_notice_tool(
        receivable_path: str,
        notice_month: str,
        output_dir: str = "",
        notice_date: str = "",
        due_date: str = "",
    ) -> str:
        """PMS付款通知书生成：读取PMS应收账务列表，按协议单位分组，
        为每个协议客户生成一份付款通知书Excel文件。
        """
        return generate_payment_notices(
            receivable_path=receivable_path,
            notice_month=notice_month,
            output_dir=output_dir,
            notice_date=notice_date or None,
            due_date=due_date or None,
        )
except ImportError:
    # 若未安装langchain，跳过Tool装饰
    payment_notice_tool = None



if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("用法: python notice_pms.py <应收账务列表.xlsx> <账期(YYYY-MM)> [通知书日期(YYYY-MM-DD)]")
        sys.exit(1)

    result = generate_payment_notices(
        receivable_path=sys.argv[1],
        notice_month=sys.argv[2],
        notice_date=sys.argv[3] if len(sys.argv) > 3 else None,
    )
    print(result)