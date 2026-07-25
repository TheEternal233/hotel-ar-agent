"""M04: 协议企业客户对账 — 生成付款通知书

从 PMS 应收账务处理列表中提取协议客户数据，按协议单位分组汇总，
生成付款通知书 (Excel) 并计算信用额度使用情况。
"""

import os, json
from datetime import datetime
from langchain.tools import tool
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

from tools.doc_parser import read_rezen

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE_DIR, "output")
CONFIG_DIR = os.path.join(BASE_DIR, "config")
os.makedirs(OUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")


def _load_credit_rules():
    path = os.path.join(CONFIG_DIR, "credit_rules.json")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _get_corp_credit(corp_name, rules):
    """根据协议单位名称获取信用额度和账期"""
    for r in rules.get("rules", []):
        if r.get("customer_type") == "企业协议":
            return r.get("credit_limit", 50000), r.get("credit_days", 60)
    return 50000, 60


@tool
def corp_recon(receivable_path: str) -> str:
    """协议企业客户对账：读取PMS应收账务处理列表，按协议单位分组汇总，
    生成付款通知书，自动检查信用额度使用情况。"""
    if not os.path.exists(receivable_path):
        return f"错误：文件不存在: {receivable_path}"

    records = read_rezen(receivable_path)

    # 筛选协议单位记录
    corp_records = [r for r in records if r.get("corp") and r["corp"].strip()]
    if not corp_records:
        return "未找到协议单位记录（corp 字段为空）"

    # 按协议单位分组
    corp_groups = {}
    for r in corp_records:
        corp = r["corp"].strip()
        if corp not in corp_groups:
            corp_groups[corp] = {"records": [], "total": 0, "count": 0}
        corp_groups[corp]["records"].append(r)
        corp_groups[corp]["total"] += r["amount"]
        corp_groups[corp]["count"] += 1

    # 加载信用规则
    rules = _load_credit_rules()
    credit_limit, credit_days = _get_corp_credit("企业协议", rules)

    # 生成报告
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUT_DIR, f"协议客户对账_{now}.xlsx")

    wb = openpyxl.Workbook()

    # Sheet 1: 付款通知书
    ws1 = wb.active
    ws1.title = "付款通知书"
    notice_hdrs = ["序号", "协议单位", "本月消费金额", "信用额度", "额度使用率", "信用账期(天)", "到期日"]
    for j, h in enumerate(notice_hdrs, 1):
        c = ws1.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    sorted_corps = sorted(corp_groups.items(), key=lambda x: -x[1]["total"])
    overdue_count = 0
    for i, (corp, data) in enumerate(sorted_corps, 2):
        utilization = round(data["total"] / credit_limit * 100, 1) if credit_limit else 0
        due_date = datetime.now().strftime("%Y-%m-%d")  # Simplified
        vals = [i - 1, corp, round(data["total"], 2), credit_limit,
                f"{utilization}%", credit_days, due_date]
        for j, v in enumerate(vals, 1):
            c = ws1.cell(row=i, column=j, value=v)
            c.border = THIN_BORDER
            if utilization > 80:
                c.fill = RED_FILL
                overdue_count += 1
            elif utilization > 50:
                c.fill = YELLOW_FILL

    # Summary row
    ri = len(sorted_corps) + 2
    total_amount = sum(d["total"] for _, d in sorted_corps)
    summary_vals = ["合计", f"{len(sorted_corps)}家", round(total_amount, 2), "", "", "", ""]
    for j, v in enumerate(summary_vals, 1):
        c = ws1.cell(row=ri, column=j, value=v)
        c.font = Font(bold=True)
        c.border = THIN_BORDER

    # Sheet 2: 消费明细
    ws2 = wb.create_sheet("消费明细")
    detail_hdrs = ["协议单位", "日期", "房号", "金额", "结账单号", "订单号", "外部订单号", "备注"]
    for j, h in enumerate(detail_hdrs, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    ri = 2
    for corp, data in sorted_corps:
        for r in data["records"]:
            date_str = r["date"].strftime("%Y-%m-%d") if r["date"] else ""
            vals = [corp, date_str, r["room"], r["amount"],
                    r["bill_no"], r["order"], r["ext_order"], r["remark"]]
            for j, v in enumerate(vals, 1):
                c = ws2.cell(row=ri, column=j, value=v)
                c.border = THIN_BORDER
            ri += 1

    wb.save(out_path)
    wb.close()

    return (
        f"协议客户对账完成\n"
        f"报告: {out_path}\n"
        f"协议单位数: {len(corp_groups)}\n"
        f"总金额: {total_amount:,.2f}\n"
        f"额度超80%预警: {overdue_count}家\n"
        f"信用额度: {credit_limit:,} 元/家"
    )


if __name__ == "__main__":
    data_dir = os.path.join(BASE_DIR, "data", "清远", "协议企业对账")
    files = [f for f in os.listdir(data_dir) if f.endswith(".xlsx")]
    if files:
        result = corp_recon.invoke({"receivable_path": os.path.join(data_dir, files[0])})
        print(result)
    else:
        print("No data files found")
