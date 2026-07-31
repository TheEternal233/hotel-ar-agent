"""信用卡对账：常量、样式与付款方式映射

只对 4 种付款方式对账：微信、支付宝、OTA卡、预付卡。
挂应收、挂房账、挂团队、OC、ENT、YFD 等付款方式一律不统计。
"""

from openpyxl.styles import Font, PatternFill, Border, Side

# ===== Excel 样式 =====
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

# ===== 对账配置 =====

# 需要对账的 4 种付款方式（规范名称，固定顺序）
RECON_PAYMENT_METHODS = ["微信", "支付宝", "OTA卡", "预付卡"]

# 不统计的付款方式关键字（命中任意一个即排除）
# 挂应收、挂房账、挂团队、OC、ENT、YFD 等
EXCLUDED_KEYWORDS = ["挂应收", "挂房账", "挂团队", "挂账", "OC", "ENT", "YFD"]

# 金额对比容差（元），小于等于该值视为金额相等
AMOUNT_TOLERANCE = 0.01


def normalize_payment(desc, source="pms"):
    """将付款描述 / 支付类型映射到 4 种规范付款方式之一。

    Args:
        desc: PMS 的「付款描述」或 POS 的「支付类型」原始字符串
        source: 数据来源标记 "pms" / "pos"（仅用于日志，映射规则一致）

    Returns:
        规范付款方式名称（微信 / 支付宝 / OTA卡 / 预付卡）；
        命中排除项或无法识别时返回 None（表示不统计）。
    """

    desc = str(desc).strip()
    if not desc:
        return None
    upper = desc.upper()

    # 1) 排除项：挂应收、挂房账、OC、ENT、YFD 等（YFD 微信也在此排除）
    for kw in EXCLUDED_KEYWORDS:
        if kw.upper() in upper:
            return None

    # 2) 映射到 4 种规范付款方式（先判断 YFD 已排除，避免「YFD 微信」误判为微信）
    if "微信" in desc or "WECHAT" in upper:
        return "微信"
    if "支付宝" in desc or "ALIPAY" in upper:
        return "支付宝"
    if "OTA" in upper:
        return "OTA卡"
    if "预付" in desc or "PREPAID" in upper:
        return "预付卡"
    return None
