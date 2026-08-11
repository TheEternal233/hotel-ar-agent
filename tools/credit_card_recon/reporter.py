"""信用卡对账：对账差异报告生成

输出一个 Excel 对账差异表格：(单sheet)
    上半部分[对账差异汇总]--每种付款方式的 PMS、POS数量、金额、差额、对平状态
    中间空三行
    下半部分[差异明细]  --未能逐笔配对的PMS短款、POS长款明细

"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tools import BASE_DIR
from tools.credit_card_recon.constants import (
    HEADER_FILL, HEADER_FONT, THIN_BORDER,
    RED_FILL, YELLOW_FILL, GREEN_FILL,
    AMOUNT_TOLERANCE,
)


def _generate_recon_report(recon_results):
    """生成对账差异报告 Excel

    Args:
        recon_results: _reconcile_channel 的结果列表

    Returns:
        str: 含各付款方式差异摘要与报告路径的文本
    """
    out_dir = os.path.join(BASE_DIR, "output", "信用卡审核")
    os.makedirs(out_dir, exist_ok=True)
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"对账差异报告_{now}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "对账差异报告"

    # ===== 上半部分: 对账差异汇总 =====

    headers = [
        "付款方式", "PMS数量", "POS数量", "数量差",
        "PMS金额", "POS金额", "金额差额",
        "数量匹配", "金额匹配", "逐笔匹配", "状态",
    ]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    for i, r in enumerate(recon_results, 2):
        count_diff = r["pms_count"] - r["bank_count"]
        row_vals = [
            r["channel"],
            r["pms_count"],
            r["bank_count"],
            count_diff,
            r["pms_total"],
            r["bank_total"],
            r["diff"],
            "是" if r["count_match"] else "否",
            "是" if r["amount_match"] else "否",
            "是" if r.get("all_matched", True) else "否",
            "对平" if r["balanced"] else "差异",
        ]
        for j, v in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN_BORDER
        # 状态列着色
        status_cell = ws.cell(row=i, column=11)
        status_cell.fill = GREEN_FILL if r["balanced"] else RED_FILL
        # 差额不为 0 的金额差额列标黄
        if abs(r["diff"]) > AMOUNT_TOLERANCE:
            ws.cell(row=i, column=7).fill = YELLOW_FILL
        if not r["count_match"]:
            ws.cell(row=i, column=4).fill = YELLOW_FILL
        # 逐笔未全部配对时标黄提醒
        if not r.get("all_matched", True):
            ws.cell(row=i, column=10).fill = YELLOW_FILL


    # 汇总部分结束行
    summary_end_row=1+len(recon_results)

    # =====中间空三行=====
    gap = 3
    detail_start=summary_end_row+gap+1

    # =====下半部分：差异明细=====
    #小节标题
    title_cell= ws.cell(row=detail_start-1,column=1,value="差异明细")
    title_cell.font=HEADER_FONT
    hdrs= ["付款方式","来源","差异类型","金额","原始数据"]

    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row=detail_start,column=j,value=h)
        c.fill=HEADER_FILL
        c.font=HEADER_FONT
        c.border=THIN_BORDER

    ri=detail_start+1
    for r in recon_results:
        # PMS短款：PMS有、POS无对应
        for um in r.get("unmatched_pms",[]):
            vals=[
                r["channel"],"PMS","PMS短款",
                um.get("amount",0),str(um.get("raw",{}))
            ]
            for j,v in enumerate(vals, 1):
                c=ws.cell(row=ri,column=j,value=v)
                c.fill=RED_FILL
                c.border=THIN_BORDER
            ri+=1
        # POS长款：POS有、PMS无
        for um in r.get("unmatched_bank",[]):
            vals=[
                r["channel"],"POS","POS长款",
                um.get("amount",0),str(um.get("raw",{}))
            ]
            for j,v in enumerate(vals, 1):
                c=ws.cell(row=ri,column=j,value=v)
                c.fill=YELLOW_FILL
                c.border=THIN_BORDER
            ri+=1

    col_widths=[12,10,10,10,60,14,14,10,10,10,8]
    for j,w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    wb.save(out_path)
    wb.close()


    #文本摘要
    summary_lines=[]
    for r in recon_results:
        flag="对平" if r["balanced"] else "差异"
        detail=(
            f"匹配{r.get('matched_count',0)}笔， "
            f"PMS短款{r.get('unmatched_pms_count',0)}/POS长款{r.get('unmatched_bank_count',0)}"
        )

        summary_lines.append(
            f"{r['channel']}:PMS数量{r['pms_count']}/POS数量{r['bank_count']}("
            f"{'数量一致' if r['count_match'] else '数量不一致'}),"
            f"PMS金额{r['pms_total']:.2f}/POS金额{r['bank_total']:.2f},"
            f"差额{r['diff']:.2f},{detail},({flag})"
        )

    return "对账完成\n"+"\n".join(summary_lines)+f"\n\n报告: {out_path}"