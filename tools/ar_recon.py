"""M02: OTA 多渠道对账工具 — 清远酒店真实数据驱动

渠道覆盖：携程客房·美团客房·飞猪·抖音·携程餐饮·美团餐饮·向蜜鸟
匹配策略：订单号精匹配 / 券号间接匹配 / 金额+日期模糊匹配
"""

import os
from datetime import datetime
from langchain.tools import tool
import openpyxl
from openpyxl.styles import Font
from enums.common_enum import HEADER_FILL, HEADER_FONT, THIN_BORDER, RED_FILL, GREEN_FILL, YELLOW_FILL, OUT_DIR, \
    BASE_DIR
from utils.common_func import _norm_orderno, _norm_amount
from tools.doc_parser import read_ota_channel, read_rezen, detect_ota_channel, OTA_CHANNEL_MAPPINGS
from utils.ar_recon_utils import _generate_ar_report, _match_ota_rezen_fnb, _generate_ar_report_fnb

os.makedirs(OUT_DIR, exist_ok=True)



CHANNEL_NAMES = {
    "携程客房": "携程客房", "携程餐饮": "携程餐饮",
    "美团客房": "美团客房", "美团餐饮": "美团餐饮",
    "飞猪": "飞猪", "抖音": "抖音", "向蜜鸟": "向蜜鸟",
}

FNB_CHANNELS = {"美团餐饮", "携程餐饮"}


def _match_ota_rezen(ota_records, rezen_records, channel_name):
    oci = OTA_CHANNEL_MAPPINGS.get(channel_name, {})
    oid_col = oci.get("order_id_col", "order_id")
    amt_col = oci.get("amount_col", "amount")

    rezen_by_ext = {}
    rezen_by_order = {}
    for i, r in enumerate(rezen_records):
        eo = _norm_orderno(r.get("ext_order", ""))
        od = _norm_orderno(r.get("order", ""))
        if eo:
            rezen_by_ext.setdefault(eo, []).append(i)
        if od:
            rezen_by_order.setdefault(od, []).append(i)

    rezen_matched = set()
    results = []
    stats = {"total_ota": len(ota_records), "total_pms": len(rezen_records),
             "match": 0, "diff": 0, "ota_only": 0, "pms_only": 0}

    # 向蜜鸟特殊：识别号匹配 + 储值卡兜底
    is_xiangminiao = channel_name == "向蜜鸟"

    for ota in ota_records:
        oid = _norm_orderno(ota.get(oid_col, ""))
        oamt = _norm_amount(ota.get(amt_col, 0))
        identify_no = _norm_orderno(ota.get("identify_no", "")) if is_xiangminiao else ""

        found = False
        ri = -1

        if oid:
            candidates = rezen_by_ext.get(oid, []) + rezen_by_order.get(oid, [])
            # 第一轮：优先找订单号匹配且金额一致的
            for ci in candidates:
                if ci in rezen_matched:
                    continue
                ramt = _norm_amount(rezen_records[ci].get("amount", 0))
                if abs(oamt - ramt) < 0.02:
                    ri = ci
                    found = True
                    break
            # 第二轮：只要订单号匹配就认定为同一订单（金额差异标记为diff）
            if not found:
                for ci in candidates:
                    if ci in rezen_matched:
                        continue
                    ri = ci
                    found = True
                    break

        # 向蜜鸟特殊策略B: 识别号(短码)匹配
        if not found and is_xiangminiao and identify_no:
            for ci in range(len(rezen_records)):
                if ci in rezen_matched:
                    continue
                rext = _norm_orderno(rezen_records[ci].get("ext_order", ""))
                rorder = _norm_orderno(rezen_records[ci].get("order", ""))
                # 识别号可能在外部订单号或订单号字段中
                if identify_no in rext or identify_no in rorder:
                    ramt = _norm_amount(rezen_records[ci].get("amount", 0))
                    if abs(oamt - ramt) < 0.02:
                        ri = ci
                        found = True
                        break

        # 向蜜鸟特殊策略C: 储值卡消费（OTA金额为0时直接匹配识别号）
        if not found and is_xiangminiao and oamt == 0 and identify_no:
            for ci in range(len(rezen_records)):
                if ci in rezen_matched:
                    continue
                rorder = _norm_orderno(rezen_records[ci].get("order", ""))
                rext = _norm_orderno(rezen_records[ci].get("ext_order", ""))
                if identify_no in rorder or identify_no in rext:
                    ri = ci
                    found = True
                    break

        if not found and not oid and not identify_no and oamt > 0:  # 必须没有订单号才允许金额兜底
            best_ci = -1
            best_diff = float("inf")
            for ci in range(len(rezen_records)):
                if ci in rezen_matched:
                    continue
                ramt = _norm_amount(rezen_records[ci].get("amount", 0))
                if ramt <= 0:
                    continue
                diff = abs(oamt - ramt)
                if diff < 5.0 and diff < best_diff:
                    best_diff = diff
                    best_ci = ci
            if best_ci >= 0:
                ri = best_ci
                found = True

        if found:
            rezen_matched.add(ri)
            ramt = _norm_amount(rezen_records[ri].get("amount", 0))
            diff = round(oamt - ramt, 2)
            if abs(diff) < 0.02:
                stats["match"] += 1
                status = "match"
            else:
                stats["diff"] += 1
                status = "diff"
            results.append({
                "status": status,
                "ota": ota,
                "pms": rezen_records[ri],
                "ota_amount": oamt,
                "pms_amount": ramt,
                "diff": diff,
                "ota_order": oid,
                "pms_ext_order": _norm_orderno(rezen_records[ri].get("ext_order", "")),
            })
        else:
            stats["ota_only"] += 1
            results.append({
                "status": "ota_only",
                "ota": ota,
                "pms": None,
                "ota_amount": oamt,
                "pms_amount": 0,
                "diff": oamt,
                "ota_order": oid,
                "pms_ext_order": "",
            })

    for i, r in enumerate(rezen_records):
        if i not in rezen_matched:
            stats["pms_only"] += 1
            results.append({
                "status": "pms_only",
                "ota": None,
                "pms": r,
                "ota_amount": 0,
                "pms_amount": _norm_amount(r.get("amount", 0)),
                "diff": _norm_amount(r.get("amount", 0)),
                "ota_order": "",
                "pms_ext_order": _norm_orderno(r.get("ext_order", "")),
            })

    return results, stats

def _generate_report(results, stats, channel_name, ota_path, pms_path):
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    cname = channel_name.replace("·", "_")
    out_path = os.path.join(OUT_DIR, f"OTA对账_{cname}_{now}.xlsx")

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "对账汇总"
    info = [
        ("渠道", channel_name),
        ("OTA文件", os.path.basename(ota_path)),
        ("PMS文件", os.path.basename(pms_path)),
        ("对账时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("OTA记录数", stats["total_ota"]),
        ("PMS记录数", stats["total_pms"]),
        ("完全匹配", stats["match"]),
        ("金额差异", stats["diff"]),
        ("仅OTA存在", stats["ota_only"]),
        ("仅PMS存在", stats["pms_only"]),
        ("", ""),
        ("OTA金额合计", round(sum(r["ota_amount"] for r in results if r["status"] != "pms_only"), 2)),
        ("PMS金额合计", round(sum(r["pms_amount"] for r in results if r["status"] != "ota_only"), 2)),
        ("净差异", round(sum(r["diff"] for r in results if r["status"] == "diff"), 2)),
    ]
    for i, (k, v) in enumerate(info, 1):
        ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
        c = ws1.cell(row=i, column=2, value=v)
        if "差异" in str(k) and isinstance(v, (int, float)) and v != 0:
            c.fill = RED_FILL

    ws2 = wb.create_sheet("差额明细")
    hdrs2 = ["OTA订单号", "PMS外部订单号", "OTA金额", "PMS金额", "差额", "状态", "房号", "PMS备注"]
    for j, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    ri = 2
    for r in results:
        if r["status"] == "match":
            continue
        pms = r.get("pms") or {}
        vals = [
            r["ota_order"],
            r["pms_ext_order"],
            r["ota_amount"],
            r["pms_amount"],
            r["diff"],
            r["status"],
            str(pms.get("room", "")),
            str(pms.get("remark", "")),
        ]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=j, value=v)
            c.border = THIN_BORDER
            if r["status"] == "diff":
                c.fill = RED_FILL
            elif r["status"] in ("ota_only", "pms_only"):
                c.fill = YELLOW_FILL
        ri += 1

    ws3 = wb.create_sheet("全额对比")
    hdrs3 = ["状态", "OTA订单号", "PMS外部订单号", "OTA金额", "PMS金额", "差额", "房号"]
    for j, h in enumerate(hdrs3, 1):
        c = ws3.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    ri = 2
    for r in results:
        pms = r.get("pms") or {}
        vals = [
            r["status"],
            r["ota_order"],
            r["pms_ext_order"],
            r["ota_amount"],
            r["pms_amount"],
            r["diff"],
            str(pms.get("room", "")),
        ]
        for j, v in enumerate(vals, 1):
            c = ws3.cell(row=ri, column=j, value=v)
            c.border = THIN_BORDER
            if r["status"] == "match":
                c.fill = GREEN_FILL
            elif r["status"] == "diff":
                c.fill = RED_FILL
            elif r["status"] in ("ota_only", "pms_only"):
                c.fill = YELLOW_FILL
        ri += 1

    wb.save(out_path)
    wb.close()
    return out_path

def batch_ota_recon(data_dir=None):
    if data_dir is None:
        data_dir = os.path.join(BASE_DIR, "data", "清远", "OTA对账")
    if not os.path.exists(data_dir):
        return f"错误：数据目录不存在: {data_dir}"

    files = [f for f in os.listdir(data_dir) if f.endswith((".xlsx", ".xls"))]
    rezen_files = [f for f in files if "rezen" in f.lower()]
    ota_files = [f for f in files if "rezen" not in f.lower()]

    all_stats = []
    all_reports = []

    # Build rezen lookup by channel name
    rezen_lookup = {}
    for rf in rezen_files:
        rf_base = os.path.splitext(rf)[0]
        rf_clean = rf_base.replace("rezen", "").replace("·", "").rstrip("0123456789")
        rezen_lookup[rf_clean] = rf

    for ota_file in ota_files:
        ota_path = os.path.join(data_dir, ota_file)
        ota_base = os.path.splitext(ota_file)[0]
        # Strip trailing digits for files like 飞猪1, 飞猪2
        import re
        ota_clean = re.sub(r'[0-9]+$', '', ota_base).strip()
        matched_rezen = None
        for rf_clean, rf in rezen_lookup.items():
            if ota_clean in rf_clean or rf_clean in ota_clean or ota_clean[:2] in rf_clean:
                matched_rezen = rf
                break
        if matched_rezen is None:
            # Try exact prefix match
            for rf in rezen_files:
                rf_base = os.path.splitext(rf)[0]
                if ota_base[:2] in rf_base and "rezen" in rf_base.lower():
                    matched_rezen = rf
                    break
        if matched_rezen is None:
            continue
        rezen_path = os.path.join(data_dir, matched_rezen)

        channel = detect_ota_channel(ota_path)
        if channel is None or channel == "rezen":
            continue

        try:
            ota_records = read_ota_channel(ota_path, channel)
            rezen_records = read_rezen(rezen_path)
        except Exception as e:
            all_stats.append(f"{channel}({ota_file}): 读取失败 - {e}")
            continue

        # 餐饮渠道走数量统计匹配，其他渠道走订单号匹配
        if channel in FNB_CHANNELS:
            results, stats = _match_ota_rezen_fnb(ota_records, rezen_records, channel)
            report_path = _generate_ar_report_fnb(results, stats, channel, ota_path, rezen_path)
        else:
            results, stats = _match_ota_rezen(ota_records, rezen_records, channel)
            report_path = _generate_report(results, stats, channel, ota_path, rezen_path)
        all_stats.append({
            "channel": channel,
            "file": ota_file,
            "stats": stats,
            "report": report_path,
        })
        all_reports.append(report_path)

    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    summary_path = os.path.join(OUT_DIR, f"OTA对账_全部汇总_{now}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "渠道汇总"
    hdrs = ["渠道", "OTA文件", "OTA记录", "PMS记录", "匹配", "差异", "仅OTA", "仅PMS", "报告文件"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    for i, s in enumerate(all_stats, 2):
        if isinstance(s, str):
            ws.cell(row=i, column=1, value=s)
            continue
        st = s["stats"]
        vals = [s["channel"], s["file"], st["total_ota"], st["total_pms"],
                st["match"], st["diff"], st["ota_only"], st["pms_only"],
                os.path.basename(s["report"])]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN_BORDER
            if j in (5, 6, 7, 8) and isinstance(v, (int, float)) and v > 0:
                c.fill = YELLOW_FILL
    wb.save(summary_path)
    wb.close()

    total_match = sum(s["stats"]["match"] for s in all_stats if isinstance(s, dict))
    total_diff = sum(s["stats"]["diff"] for s in all_stats if isinstance(s, dict))
    total_ota_only = sum(s["stats"]["ota_only"] for s in all_stats if isinstance(s, dict))
    total_pms_only = sum(s["stats"]["pms_only"] for s in all_stats if isinstance(s, dict))

    channel_lines = []
    for s in all_stats:
        if isinstance(s, str):
            channel_lines.append(f"  {s}")
            continue
        st = s["stats"]
        channel_lines.append(
            f"  {s['channel']}: 匹配{st['match']} 差异{st['diff']} 仅OTA{st['ota_only']} 仅PMS{st['pms_only']}"
        )

    return (
        f"OTA批量对账完成\n"
        f"渠道数: {len([s for s in all_stats if isinstance(s, dict)])}\n"
        f"匹配: {total_match}  差异: {total_diff}  仅OTA: {total_ota_only}  仅PMS: {total_pms_only}\n"
        f"汇总报告: {summary_path}\n\n"
        f"各渠道明细:\n" + "\n".join(channel_lines)
    )

@tool
def ar_recon(ota_path: str = "", pms_path: str = "", channel: str = "") -> str:
    """OTA对账：输入OTA导出和PMS(rezen)两个xlsx路径及渠道名，自动匹配并输出差额报告。

    渠道可选: 携程客房, 携程餐饮, 美团客房, 美团餐饮, 飞猪, 抖音, 向蜜鸟
    留空则自动检测渠道。
    不传参数时自动执行 data/清远/OTA对账 目录下的批量对账。
    """
    if not ota_path and not pms_path:
        return batch_ota_recon()

    if not ota_path or not pms_path:
        return "错误：请同时提供 ota_path 和 pms_path，或都不提供进行批量对账"

    if not os.path.exists(ota_path):
        return f"错误：OTA文件不存在: {ota_path}"
    if not os.path.exists(pms_path):
        return f"错误：PMS文件不存在: {pms_path}"

    if not channel:
        channel = detect_ota_channel(ota_path)
        if channel is None or channel == "rezen":
            return f"无法自动检测渠道，请手动指定 channel 参数。可用: {list(OTA_CHANNEL_MAPPINGS.keys())}"

    try:
        ota_records = read_ota_channel(ota_path, channel)
        rezen_records = read_rezen(pms_path)
    except Exception as e:
        return f"读取文件失败: {e}"

    # 餐饮渠道按金额数量统计匹配，其他渠道按订单号匹配
    if channel in FNB_CHANNELS:
        results, stats = _match_ota_rezen_fnb(ota_records, rezen_records, channel)
        report_path = _generate_ar_report_fnb(results, stats, channel, ota_path, pms_path)
    else:
        results, stats = _match_ota_rezen(ota_records, rezen_records, channel)
        report_path = _generate_ar_report(results, stats, channel, ota_path, pms_path)

    return (
        f"OTA对账完成 [{channel}]\n"
        f"报告: {report_path}\n"
        f"OTA记录: {stats['total_ota']}  PMS记录: {stats['total_pms']}\n"
        f"匹配: {stats['match']}  差异: {stats['diff']}  仅OTA: {stats['ota_only']}  仅PMS: {stats['pms_only']}"
    )

if __name__ == "__main__":
    result = batch_ota_recon()
    print(result)