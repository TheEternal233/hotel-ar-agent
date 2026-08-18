import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from langchain.tools import tool

from tools.ar_recon.constants import FNB_CHANNELS, PMS_MARKER, SUPPORTED_EXTS
from enums.common_enum import OUT_DIR
from tools.ar_recon.matcher import _match_ota_rezen, _match_ota_rezen_fnb, match_xiangminiao
from tools.ar_recon.batch_runner import batch_ota_recon
from tools.ar_recon.report_generator import _generate_ar_report_fnb, _generate_report, _generate_ar_report, \
    _generate_ar_report_a
from tools.doc_parser import read_ota_channel, read_rezen, detect_ota_channel_fast, read_sheet
from utils.ar_recon_utils import read_xiangminiao

os.makedirs(OUT_DIR, exist_ok=True)

# 批量对账最大并行工作线程数
BATCH_MAX_WORKERS = min(os.cpu_count() or 4, 4)

def _cleanup_uploads(paths):
    for f in paths:
        try:
            os.remove(f)
        except OSError:
            pass

def _process_single_channel(ota_path, pms_path, channel):
    if channel == "向蜜鸟":
        target = ota_path if os.path.exists(ota_path) else pms_path
        ota_records, card_records, rezen_records = read_xiangminiao(target)
        results, stats = match_xiangminiao(ota_records, rezen_records, card_records)
        # 向蜜鸟单文件，预加载 wb 避免 _init_workbook 重复读取
        xmn_wb = openpyxl.load_workbook(ota_path, data_only=False)
        try:
            report_path = _generate_ar_report(
                results, stats, channel, ota_path, pms_path,
                ota_wb=xmn_wb, pms_wb=xmn_wb,
            )
        finally:
            xmn_wb.close()
    elif channel in FNB_CHANNELS:
        # 只加载一次 data_only=False，同时用于读取数据和复制源文件样式到报告
        ota_wb = openpyxl.load_workbook(ota_path, data_only=False)
        pms_wb = openpyxl.load_workbook(pms_path, data_only=False)
        try:
            ota_records = read_ota_channel(ota_path, channel, wb=ota_wb)
            rezen_records = read_rezen(pms_path, wb=pms_wb)
            results, stats = _match_ota_rezen_fnb(ota_records, rezen_records, channel)
            report_path = _generate_ar_report_fnb(
                results, stats, channel, ota_path, pms_path,
                ota_wb=ota_wb, pms_wb=pms_wb,
            )
        finally:
            ota_wb.close()
            pms_wb.close()
    else:
        # 只加载一次 data_only=False，同时用于读取数据和复制源文件样式到报告
        ota_wb = openpyxl.load_workbook(ota_path, data_only=False)
        pms_wb = openpyxl.load_workbook(pms_path, data_only=False)
        try:
            ota_records = read_ota_channel(ota_path, channel, wb=ota_wb)
            rezen_records = read_rezen(pms_path, wb=pms_wb)
            results, stats = _match_ota_rezen(ota_records, rezen_records, channel)
            # A类报告需要PMS原始表头+行数据，从已加载的 pms_wb 直接读取避免重复加载
            pms_headers, pms_raw_rows = read_sheet(pms_path, wb=pms_wb)
            report_path = _generate_ar_report_a(
                results, stats, channel, ota_path, pms_path,
                pms_headers=pms_headers, pms_raw_rows=pms_raw_rows,
                ota_wb=ota_wb, pms_wb=pms_wb,
            )
        finally:
            ota_wb.close()
            pms_wb.close()
    return results, stats, report_path

def _build_upload_pairs(upload_dir):
    files = sorted([f for f in os.listdir(upload_dir) if f.endswith(SUPPORTED_EXTS)])
    if not files:
        return [], []

    rezen_files = [f for f in files if PMS_MARKER in f.lower()]
    ota_files = [f for f in files if PMS_MARKER not in f.lower()]

    rezen_lookup = {}
    for rf in rezen_files:
        rf_base = os.path.splitext(rf)[0]
        rf_clean = rf_base.replace(PMS_MARKER, "").replace("·", "").rstrip("0123456789")
        rezen_lookup[rf_clean] = rf

    pairs = []

    for ota_file in ota_files:
        ota_path = os.path.join(upload_dir, ota_file)
        channel = detect_ota_channel_fast(ota_path)
        if channel is None or channel == PMS_MARKER:
            continue

        if channel == "向蜜鸟":
            pairs.append((ota_path, ota_path, channel))
            continue

        ota_base = os.path.splitext(ota_file)[0]
        ota_clean = re.sub(r'[0-9]+$', '', ota_base).strip()
        matched_rezen = None
        for rf_clean, rf in rezen_lookup.items():
            if ota_clean in rf_clean or rf_clean in ota_clean or channel in rf_clean:
                matched_rezen = rf
                break
        if matched_rezen is None:
            for rf in rezen_files:
                rf_base = os.path.splitext(rf)[0]
                if channel in rf_base and PMS_MARKER in rf_base.lower():
                    matched_rezen = rf
                    break
        if matched_rezen is None:
            continue
        rezen_path = os.path.join(upload_dir, matched_rezen)
        pairs.append((ota_path, rezen_path, channel))

    unused_rezen = [rf for rf in rezen_files if not any(rf == os.path.basename(p[1]) for p in pairs)]
    return pairs, unused_rezen

@tool
def ar_recon(ota_path: str = "", pms_path: str = "", channel: str = "") -> str:
    """OTA对账：输入OTA导出和PMS(rezen)两个xlsx路径及渠道名，自动匹配并输出差额报告。

    渠道可选: 携程客房, 携程餐饮, 美团客房, 美团餐饮, 飞猪, 抖音, 向蜜鸟
    留空则自动检测渠道。
    不传参数时自动扫描 uploads/ 目录下的文件，按文件名自动配对多平台。
    """
    _cleanup = set()
    if not ota_path and not pms_path:
        from enums.common_enum import BASE_DIR
        upload_dir = os.path.join(BASE_DIR, "uploads")
        if not os.path.exists(upload_dir):
            return "请上传文件后再对账，或提供 ota_path 和 pms_path。"

        pairs, unused_rezen = _build_upload_pairs(upload_dir)
        if not pairs:
            files = os.listdir(upload_dir)
            return f"uploads/ 目录文件: {', '.join(files)}。未找到可配对的OTA和PMS文件。"

        # 使用线程池并行处理各渠道（渠道之间无依赖）
        all_results = []
        with ThreadPoolExecutor(max_workers=BATCH_MAX_WORKERS) as executor:
            future_to_pair = {
                executor.submit(_process_single_channel, op, pp, ch): (op, pp, ch)
                for op, pp, ch in pairs
            }
            for future in as_completed(future_to_pair):
                op, pp, ch = future_to_pair[future]
                _cleanup.add(op)
                _cleanup.add(pp)
                try:
                    results, stats, report_path = future.result()
                    all_results.append({
                        "channel": ch,
                        "stats": stats,
                        "report": report_path,
                    })
                except Exception as e:
                    all_results.append({
                        "channel": ch,
                        "error": str(e),
                    })

        _cleanup_uploads(_cleanup)

        channel_lines = []
        total_match = total_diff = total_ota_only = total_pms_only = 0
        for r in all_results:
            if "error" in r:
                channel_lines.append(f"  {r['channel']}: 读取失败 - {r['error']}")
                continue
            st = r["stats"]
            total_match += st["match"]
            total_diff += st["diff"]
            total_ota_only += st["ota_only"]
            total_pms_only += st["pms_only"]
            channel_lines.append(
                f"  {r['channel']}: 匹配{st['match']} 差异{st['diff']} 仅OTA{st['ota_only']} 仅PMS{st['pms_only']}"
            )

        return (
            f"OTA批量对账完成\n"
            f"渠道数: {len([r for r in all_results if 'error' not in r])}\n"
            f"匹配: {total_match}  差异: {total_diff}  仅OTA: {total_ota_only}  仅PMS: {total_pms_only}\n\n"
            f"各渠道明细:\n" + "\n".join(channel_lines)
        )

    if not ota_path or not pms_path:
        return "错误：请同时提供 ota_path 和 pms_path，或都不提供进行批量对账"

    if not os.path.exists(ota_path):
        return f"错误：OTA文件不存在: {ota_path}"
    if not os.path.exists(pms_path):
        return f"错误：PMS文件不存在: {pms_path}"

    from enums.common_enum import BASE_DIR
    upload_dir = os.path.join(BASE_DIR, "uploads")
    if not _cleanup:
        for p in (ota_path, pms_path):
            if os.path.abspath(p).startswith(os.path.abspath(upload_dir)):
                _cleanup.add(p)

    if not channel:
        channel = detect_ota_channel_fast(ota_path)
        if channel is None or channel == PMS_MARKER:
            return f"无法自动检测渠道，请手动指定 channel 参数。"

    try:
        results, stats, report_path = _process_single_channel(ota_path, pms_path, channel)
    except Exception as e:
        _cleanup_uploads(_cleanup)
        return f"读取文件失败: {e}"

    _cleanup_uploads(_cleanup)

    return (
        f"OTA对账完成 [{channel}]\n"
        f"报告: {report_path}\n"
        f"OTA记录: {stats['total_ota']}  PMS记录: {stats['total_pms']}\n"
        f"匹配: {stats['match']}  差异: {stats['diff']}  仅OTA: {stats['ota_only']}  仅PMS: {stats['pms_only']}"
    )

__all__ = ["ar_recon"]