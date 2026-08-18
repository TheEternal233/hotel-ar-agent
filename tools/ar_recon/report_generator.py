import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font

from enums.common_enum import (
    OUT_DIR, HEADER_FILL, HEADER_FONT, THIN_BORDER,
    RED_FILL, GREEN_FILL, YELLOW_FILL,
)
from tools.doc_parser import read_sheet
from utils.ar_recon_utils import _copy_sheet_to_wb
from tools.ar_recon.constants import (
    REPORT_FILENAME_FMT, XIANGMINIAO_OTA_SHEET,
    SHEET_SUMMARY, SHEET_DIFF, SHEET_FULL,
    STATUS_MATCH, STATUS_DIFF, STATUS_OTA_ONLY, STATUS_PMS_ONLY,
    STD_DIFF_HDRS, STD_FULL_HDRS, FNB_DIFF_HDRS, FNB_FULL_HDRS,
    AR_DIFF_HDRS, STATUS_ORDER,
    OTA_RECON_DIR, OTA_PREFIX, PMS_PREFIX,
)

def _make_out_path(channel_name):
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    cname = channel_name.replace("·", "_")
    ota_dir = os.path.join(OUT_DIR, OTA_RECON_DIR)
    os.makedirs(ota_dir, exist_ok=True)
    return os.path.join(ota_dir, REPORT_FILENAME_FMT.format(cname, now))


def _init_workbook(ar_mode=False, ota_path=None, pms_path=None,
                     ota_wb=None, pms_wb=None):
    """初始化 Workbook；AR 模式下复制源文件 sheet

    支持传入已加载的 workbook 对象(ota_wb/pms_wb)，避免重复从磁盘加载。
    调用方负责关闭传入的 workbook。
    """
    wb = openpyxl.Workbook()
    if not ar_mode:
        return wb

    wb.remove(wb.active)
    if ota_path == pms_path:
        src_wb = ota_wb or openpyxl.load_workbook(ota_path, data_only=False)
        try:
            for ws in src_wb.worksheets:
                new_title = OTA_PREFIX if ws.title == XIANGMINIAO_OTA_SHEET else None
                _copy_sheet_to_wb(ws, wb, title=new_title)
        finally:
            if ota_wb is None:
                src_wb.close()
    else:
        for path, prefix, cached_wb in (
            (pms_path, PMS_PREFIX, pms_wb),
            (ota_path, OTA_PREFIX, ota_wb),
        ):
            src_wb = cached_wb or openpyxl.load_workbook(path, data_only=False)
            try:
                for idx, ws in enumerate(src_wb.worksheets):
                    _copy_sheet_to_wb(ws, wb, title=prefix if idx == 0 else None)
            finally:
                if cached_wb is None:
                    src_wb.close()
    return wb


def _write_summary_block(ws, start_row, info_items, red_check_fn=None):
    """写入汇总信息块，返回下一行号"""
    row = start_row
    for k, v in info_items:
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=v)
        if red_check_fn and red_check_fn(k, v):
            c.fill = RED_FILL
        row += 1
    return row


def _write_section_title(ws, row, title):
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
    return row + 1


def _write_headers(ws, row, headers):
    ws.append(headers)
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=row, column=j)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER
    return row + 1


def _apply_status_color(cell, status):
    if status == STATUS_MATCH:
        cell.fill = GREEN_FILL
    elif status == STATUS_DIFF:
        cell.fill = RED_FILL
    elif status in (STATUS_OTA_ONLY,STATUS_PMS_ONLY):
        cell.fill = YELLOW_FILL


def _write_data_rows(ws, start_row, results, cols_fn, skip_match=False):
    """通用数据行写入，按状态着色，返回下一行号"""
    row = start_row
    written = []  # 记录实际写入的 (row_index, status)
    for r in results:
        if skip_match and r["status"] == STATUS_MATCH:
            continue
        ws.append(cols_fn(r))
        written.append((row, r["status"]))
        row += 1
    # 批量设置样式
    for r_idx, status in written:
        for c in ws[r_idx]:
            c.border = THIN_BORDER
            _apply_status_color(c, status)
    return row


def _save_and_close(wb, out_path):
    wb.save(out_path)
    wb.close()
    return out_path






def _std_diff_cols(r):
    pms = r.get("pms") or {}
    return [
        r["ota_order"], r["pms_ext_order"], r["ota_amount"],
        r["pms_amount"], r["diff"], r["status"],
        str(pms.get("room", "")), str(pms.get("remark", "")),
    ]


def _std_full_cols(r):
    pms = r.get("pms") or {}
    return [
        r["status"], r["ota_order"], r["pms_ext_order"],
        r["ota_amount"], r["pms_amount"], r["diff"],
        str(pms.get("room", "")),
    ]


def _build_std_info(results, stats, channel_name, ota_path, pms_path):
    return [
        ("渠道", channel_name),
        ("OTA文件", os.path.basename(ota_path)),
        ("PMS文件", os.path.basename(pms_path)),
        ("对账时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("OTA记录数", stats["total_ota"]),
        ("PMS记录数", stats["total_pms"]),
        ("完全匹配", stats["match"]),
        ("金额差异", stats["diff"]),
        ("仅OTA存在", stats["ota_only"]),
        ("仅PMS存在", stats["pms_only"]),
        ("", ""),
        ("OTA金额合计", round(sum(r["ota_amount"] for r in results if r["status"] != STATUS_PMS_ONLY), 2)),
        ("PMS金额合计", round(sum(r["pms_amount"] for r in results if r["status"] != STATUS_OTA_ONLY), 2)),
        ("净差异", round(sum(r["diff"] for r in results if r["status"] == STATUS_DIFF), 2)),
    ]


def _std_red_check(k, v):
    return "差异" in str(k) and isinstance(v, (int, float)) and v != 0




def _fnb_diff_cols(r):
    return [
        r["price"], r["ota_count"], r["pms_count"], r["diff_count"],
        r["ota_amount_total"], r["pms_amount_total"], r["diff_amount"],
        r["status"], r.get("ota_vouchers", ""), r.get("pms_bills", ""),
    ]


def _fnb_full_cols(r):
    return [
        r["status"], r["price"], r["ota_count"], r["pms_count"], r["diff_count"],
        r["ota_amount_total"], r["pms_amount_total"], r["diff_amount"],
        r.get("ota_vouchers", ""), r.get("pms_bills", ""),
    ]


def _build_fnb_info(results, stats, channel_name, ota_path, pms_path):
    return [
        ("渠道", channel_name),
        ("OTA文件", os.path.basename(ota_path)),
        ("PMS文件", os.path.basename(pms_path)),
        ("对账时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("OTA记录数", stats["total_ota"]),
        ("PMS记录数", stats["total_pms"]),
        ("数量完全匹配", stats["match"]),
        ("数量差异", stats["diff"]),
        ("仅OTA存在", stats["ota_only"]),
        ("仅PMS存在", stats["pms_only"]),
        ("", ""),
        ("OTA金额合计", stats.get("ota_amount_total", 0)),
        ("PMS金额合计", stats.get("pms_amount_total", 0)),
        ("净金额差异", round(sum(r["diff_amount"] for r in results if r["status"] == STATUS_DIFF), 2)),
    ]


def _fnb_red_check(k, v):
    return ("差异" in str(k) or "差额" in str(k)) and isinstance(v, (int, float)) and v != 0



def _generate_report(results, stats, channel_name, ota_path, pms_path):
    out_path = _make_out_path(channel_name)
    wb = _init_workbook()

    ws1 = wb.active
    ws1.title = SHEET_SUMMARY
    _write_summary_block(ws1, 1, _build_std_info(results, stats, channel_name, ota_path, pms_path), _std_red_check)

    ws2 = wb.create_sheet(SHEET_DIFF)
    _write_headers(ws2, 1, STD_DIFF_HDRS)
    _write_data_rows(ws2, 2, results, _std_diff_cols, skip_match=True)

    ws3 = wb.create_sheet(SHEET_FULL)
    _write_headers(ws3, 1, STD_FULL_HDRS)
    _write_data_rows(ws3, 2, results, _std_full_cols, skip_match=False)

    return _save_and_close(wb, out_path)


def _generate_ar_report(results, stats, channel_name, ota_path, pms_path,
                        ota_wb=None, pms_wb=None):
    """生成AR报告，支持传入已加载的 workbook 避免重复读取源文件。"""
    out_path = _make_out_path(channel_name)
    wb = _init_workbook(ar_mode=True, ota_path=ota_path, pms_path=pms_path,
                        ota_wb=ota_wb, pms_wb=pms_wb)

    _STATUS_ORDER = {STATUS_MATCH: 0, STATUS_DIFF: 1, STATUS_PMS_ONLY: 2, STATUS_OTA_ONLY: 3}
    results = sorted(results, key=lambda r: _STATUS_ORDER.get(r["status"], 9))

    ws = wb.create_sheet(SHEET_SUMMARY)
    row = _write_summary_block(ws, 1, _build_std_info(results, stats, channel_name, ota_path, pms_path), _std_red_check)

    row += 1
    row = _write_section_title(ws, row, "差额明细（仅展示未匹配/有差异的记录）")
    row = _write_headers(ws, row, STD_DIFF_HDRS)
    row = _write_data_rows(ws, row, results, _std_diff_cols, skip_match=True)

    row += 1
    row = _write_section_title(ws, row, "全额对比（含匹配记录）")
    row = _write_headers(ws, row, STD_FULL_HDRS)
    row = _write_data_rows(ws, row, results, _std_full_cols, skip_match=False)

    return _save_and_close(wb, out_path)


def _generate_ar_report_fnb(results, stats, channel_name, ota_path, pms_path,
                            ota_wb=None, pms_wb=None):
    """生成F&B渠道AR报告，支持传入已加载的 workbook 避免重复读取源文件。"""
    out_path = _make_out_path(channel_name)
    wb = _init_workbook(ar_mode=True, ota_path=ota_path, pms_path=pms_path,
                        ota_wb=ota_wb, pms_wb=pms_wb)

    _STATUS_ORDER = {STATUS_MATCH: 0, STATUS_DIFF: 1, STATUS_PMS_ONLY: 2, STATUS_OTA_ONLY: 3}
    results = sorted(results, key=lambda r: _STATUS_ORDER.get(r["status"], 9))

    ws = wb.create_sheet(SHEET_SUMMARY)
    row = _write_summary_block(ws, 1, _build_fnb_info(results, stats, channel_name, ota_path, pms_path), _fnb_red_check)

    row += 1
    row = _write_section_title(ws, row, "差额明细（按金额分组，仅展示数量不一致的记录）")
    row = _write_headers(ws, row, FNB_DIFF_HDRS)
    row = _write_data_rows(ws, row, results, _fnb_diff_cols, skip_match=True)

    row += 1
    row = _write_section_title(ws, row, "全额对比（含数量匹配记录）")
    row = _write_headers(ws, row, FNB_FULL_HDRS)
    row = _write_data_rows(ws, row, results, _fnb_full_cols, skip_match=False)

    return _save_and_close(wb, out_path)


def _generate_ar_report_a(results, stats, channel_name, ota_path, pms_path,
                           pms_headers=None, pms_raw_rows=None):
    """A类渠道报告，前两个为PMS，OTA源文件，第三为PMS原表+差异字段，ota_only单独列出

    支持传入已读取的 pms_headers 和 pms_raw_rows，避免重复读取 PMS 文件。
    """
    out_path = _make_out_path(channel_name)

    # 若调用方已提供解析后的数据，则复用；否则回退到重新读取
    if pms_headers is None or pms_raw_rows is None:
        pms_headers, pms_raw_rows = read_sheet(pms_path)

    wb = _init_workbook(ar_mode=True, ota_path=ota_path, pms_path=pms_path)

    ws = wb.create_sheet(SHEET_SUMMARY)

    while pms_headers and pms_headers[-1] == "":
        pms_headers.pop()

    bill_to_row = {}
    for row in pms_raw_rows:
        bid = str(row.get("账单号", "") or "")
        if bid:
            bill_to_row[bid] = row

    bill_to_result = {}
    ota_only_results = []
    for r in results:
        if r["status"] == STATUS_OTA_ONLY:
            ota_only_results.append(r)
            continue
        pms = r.get("pms")
        if pms:
            bid = pms.get("bill_id", "")
            if bid:
                bill_to_result[bid] = r

    all_hdrs = pms_headers + AR_DIFF_HDRS

    n_pms = len(pms_headers)
    _write_headers(ws, 1, all_hdrs)

    _STATUS_ORDER = {STATUS_MATCH: 0, STATUS_DIFF: 1, STATUS_PMS_ONLY: 2, STATUS_OTA_ONLY: 3}
    sorted_results = sorted(results, key=lambda r: _STATUS_ORDER.get(r["status"], 9))
    row_idx = 2
    written = []
    for r in sorted_results:
        pms = r.get("pms")
        raw_row = {}
        if pms:
            bid = pms.get("bill_id", "")
            raw_row = bill_to_row.get(bid, {})

        row_data = [raw_row.get(h) for h in pms_headers]
        row_data += [r["status"], r.get("ota_order", ""), r.get("ota_amount", 0), r.get("diff", 0), ""]
        ws.append(row_data)
        written.append((row_idx, r["status"]))
        row_idx += 1

    # 批量设置样式
    for r_idx, status in written:
        for c in ws[r_idx]:
            c.border = THIN_BORDER
            _apply_status_color(c, status)

    return _save_and_close(wb, out_path)