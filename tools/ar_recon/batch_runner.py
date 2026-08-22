import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import openpyxl

from tools.ar_recon import _match_ota_rezen_fnb, _match_ota_rezen, FNB_CHANNELS, match_xiangminiao
from enums.common_enum import YELLOW_FILL, THIN_BORDER, HEADER_FILL, BASE_DIR, OUT_DIR, HEADER_FONT
from tools.ar_recon.report_generator import _generate_report, _generate_ar_report_fnb, _generate_ar_report, \
    _generate_ar_report_a
from tools.doc_parser import detect_ota_channel_fast, read_ota_channel, read_rezen, read_sheet
from utils.ar_recon_utils import read_xiangminiao
from tools.ar_recon.constants import (
    SUPPORTED_EXTS, PMS_MARKER, DEFAULT_DATA_SUBDIR, SUMMARY_FILENAME_FMT,
    SUMMARY_SHEET_NAME, SUMMARY_HEADERS, HIGHLIGHT_COLS, PREFIX_MATCH_LEN,
    OTA_RECON_DIR,
)

# 批量对账最大并行工作线程数
# Excel 读取是 I/O 密集型任务，线程数可适当提高
BATCH_MAX_WORKERS = min((os.cpu_count() or 4) * 2, 8)

def _process_one_channel(data_dir, ota_file, rezen_lookup, rezen_files):
    """处理单个渠道（用于线程池并行）"""
    ota_path = os.path.join(data_dir, ota_file)
    ota_base = os.path.splitext(ota_file)[0]
    channel = detect_ota_channel_fast(ota_path)
    if channel is None or channel == PMS_MARKER:
        return None

    if channel == "向蜜鸟":
        xmn_wb = None
        try:
            xmn_wb = openpyxl.load_workbook(ota_path, data_only=False)
            ota_records, card_records, rezen_records = read_xiangminiao(ota_path, wb=xmn_wb)
        except Exception as e:
            if xmn_wb is not None:
                xmn_wb.close()
            return f"向蜜鸟({ota_file}): 读取失败 - {e}"
        results, stats = match_xiangminiao(ota_records, rezen_records, card_records)
        try:
            report_path = _generate_ar_report(
                results, stats, channel, ota_path, ota_path,
                ota_wb=xmn_wb, pms_wb=xmn_wb,
            )
        finally:
            xmn_wb.close()
        return {
            "channel": channel,
            "file": ota_file,
            "stats": stats,
            "report": report_path,
        }

    ota_clean = re.sub(r'[0-9]+$', '', ota_base).strip()
    matched_rezen = None

    # 第一轮：精确匹配（ota_clean 和 rf_clean 互相包含）
    for rf_clean, rf in rezen_lookup.items():
        if ota_clean == rf_clean or ota_clean in rf_clean or rf_clean in ota_clean:
            matched_rezen = rf
            break

    # 第二轮：前缀匹配（要求 ota_base 前3个字符在 rezen 文件名中，且不是仅前2个字的模糊匹配）
    if matched_rezen is None:
        for rf in rezen_files:
            rf_base = os.path.splitext(rf)[0]
            if ota_base[:PREFIX_MATCH_LEN] in rf_base and PMS_MARKER in rf_base.lower():
                matched_rezen = rf
                break
    if matched_rezen is None:
        return None
    rezen_path = os.path.join(data_dir, matched_rezen)

    try:
        # 只加载一次 data_only=False，同时用于读取数据和复制源文件样式到报告
        ota_wb = openpyxl.load_workbook(ota_path, data_only=False)
        pms_wb = openpyxl.load_workbook(rezen_path, data_only=False)
        ota_records = read_ota_channel(ota_path, channel, wb=ota_wb)
        rezen_records = read_rezen(rezen_path, wb=pms_wb)
    except Exception as e:
        return f"{channel}({ota_file}): 读取失败 - {e}"

    if channel in FNB_CHANNELS:
        results, stats = _match_ota_rezen_fnb(ota_records, rezen_records, channel)
    else:
        results, stats = _match_ota_rezen(ota_records, rezen_records, channel)

    if channel in FNB_CHANNELS:
        try:
            report_path = _generate_ar_report_fnb(
                results, stats, channel, ota_path, rezen_path,
                ota_wb=ota_wb, pms_wb=pms_wb,
            )
        finally:
            ota_wb.close()
            pms_wb.close()
    else:
        # A类报告需要PMS原始表头+行数据，从已加载的 pms_wb 直接读取避免重复加载
        pms_headers, pms_raw_rows = read_sheet(rezen_path, wb=pms_wb)
        try:
            report_path = _generate_ar_report_a(
                results, stats, channel, ota_path, rezen_path,
                pms_headers=pms_headers, pms_raw_rows=pms_raw_rows,
                ota_wb=ota_wb, pms_wb=pms_wb,
            )
        finally:
            ota_wb.close()
            pms_wb.close()
    return {
        "channel": channel,
        "file": ota_file,
        "stats": stats,
        "report": report_path,
    }


def batch_ota_recon(data_dir=None, cleanup: bool = False):
    if data_dir is None:
        data_dir = os.path.join(BASE_DIR, *DEFAULT_DATA_SUBDIR)
    if not os.path.exists(data_dir):
        return f"错误：数据目录不存在: {data_dir}"

    files = [f for f in os.listdir(data_dir) if f.endswith(SUPPORTED_EXTS)]
    rezen_files = [f for f in files if PMS_MARKER in f.lower()]
    ota_files = [f for f in files if PMS_MARKER not in f.lower()]

    all_stats = []
    all_reports = []
    used_files = set()

    # Build rezen lookup by channel name
    rezen_lookup = {}
    for rf in rezen_files:
        rf_base = os.path.splitext(rf)[0]
        rf_clean = rf_base.replace(PMS_MARKER, "").replace("·", "").rstrip("0123456789")
        rezen_lookup[rf_clean] = rf

    # 使用线程池并行处理各渠道（渠道之间无依赖）
    with ThreadPoolExecutor(max_workers=BATCH_MAX_WORKERS) as executor:
        future_to_file = {
            executor.submit(_process_one_channel, data_dir, ota_file, rezen_lookup, rezen_files): ota_file
            for ota_file in ota_files
        }
        for future in as_completed(future_to_file):
            result = future.result()
            if result is None:
                continue
            if isinstance(result, str):
                all_stats.append(result)
            else:
                all_stats.append(result)
                all_reports.append(result["report"])
                used_files.add(result["file"])
                # 记录匹配的 rezen 文件（使用与上面相同的精确匹配逻辑）
                ota_base = os.path.splitext(result["file"])[0]
                ota_clean = re.sub(r'[0-9]+$', '', ota_base).strip()
                for rf_clean, rf in rezen_lookup.items():
                    if ota_clean == rf_clean or ota_clean in rf_clean or rf_clean in ota_clean:
                        used_files.add(rf)
                        break

    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    ota_dir = os.path.join(OUT_DIR, OTA_RECON_DIR)
    os.makedirs(ota_dir, exist_ok=True)
    summary_path = os.path.join(ota_dir, SUMMARY_FILENAME_FMT.format(now))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SUMMARY_SHEET_NAME
    hdrs = SUMMARY_HEADERS
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    for i, s in enumerate(all_stats, 2):
        if isinstance(s, str):
            ws.cell(row=i, column=1, value=s)
            continue
        st = s["stats"]
        vals = [s["channel"], s["file"], st["total_ota"], st["total_pms"],
                st["match"], st["diff"], st["ota_only"], st["pms_only"],
                os.path.basename(s["report"])]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN_BORDER
            if j in HIGHLIGHT_COLS and isinstance(v, (int, float)) and v > 0:
                c.fill = YELLOW_FILL
    wb.save(summary_path)
    wb.close()

    # 清理已使用的源文件
    deleted = []
    if cleanup:
        for f in used_files:
            fp = os.path.join(data_dir, f)
            try:
                if os.path.exists(fp):
                    os.remove(fp)
                    deleted.append(f)
            except OSError:
                pass

    total_match = sum(s["stats"]["match"] for s in all_stats if isinstance(s, dict))
    total_diff = sum(s["stats"]["diff"] for s in all_stats if isinstance(s, dict))
    total_ota_only = sum(s["stats"]["ota_only"] for s in all_stats if isinstance(s, dict))
    total_pms_only = sum(s["stats"]["pms_only"] for s in all_stats if isinstance(s, dict))

    channel_lines = []
    for s in all_stats:
        if isinstance(s, str):
            channel_lines.append(f"  {s}")
            continue
        st = s["stats"]
        channel_lines.append(
            f"  {s['channel']}: 匹配{st['match']} 差异{st['diff']} 仅OTA{st['ota_only']} 仅PMS{st['pms_only']}"
        )

    result_msg = (
        f"OTA批量对账完成\n"
        f"渠道数: {len([s for s in all_stats if isinstance(s, dict)])}\n"
        f"匹配: {total_match}  差异: {total_diff}  仅OTA: {total_ota_only}  仅PMS: {total_pms_only}\n"
        f"汇总报告: {summary_path}\n\n"
        f"各渠道明细:\n" + "\n".join(channel_lines)
    )
    if cleanup and deleted:
        result_msg += f"\n\n已清理文件: {', '.join(deleted)}"
    return result_msg