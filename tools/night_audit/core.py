"""night_audit 核心：汇总 OTA 与信用卡对账结果到单一 Excel"""


from typing import List

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _read_sheet_rows(wb_path: str, sheet_name: str) -> List[List]:
    """读取指定 sheet 的所有行数据（含表头）—— 纯值版本，保留给旧代码用"""
    wb = load_workbook(wb_path, data_only=True)
    try:
        ws = wb[sheet_name]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_first_sheet_rows(wb_path: str) -> List[List]:
    """读取第一个 sheet 的所有行数据 —— 纯值版本"""
    wb = load_workbook(wb_path, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

#  带样式读取
def _cell_info(cell):
    """提取单元格的值和关键样式（填充色、字体颜色、粗体）"""
    info = {"value": cell.value}

    # 填充色（只保留非默认的 solid 填充）
    if cell.fill and cell.fill.fill_type == "solid":
        rgb = cell.fill.start_color.rgb
        if isinstance(rgb, str) and rgb not in (None, "00000000", "00FFFFFF", "FFFFFFFF"):
            info["fill"] = PatternFill(start_color=rgb[-6:], end_color=rgb[-6:], fill_type="solid")

    # 字体颜色 & 粗体
    font_color = None
    if cell.font and cell.font.color and cell.font.color.rgb:
        rgb = cell.font.color.rgb
        if isinstance(rgb, str) and rgb not in (None, "00000000", "00FFFFFF", "FFFFFFFF"):
            font_color = rgb[-6:]
    if font_color or (cell.font and cell.font.bold):
        info["font"] = Font(color=font_color, bold=cell.font.bold if cell.font else False)

    return info


def _read_sheet_rows_with_style(wb_path: str, sheet_name: str) -> List[List[dict]]:
    """读取指定 sheet，返回 [[{'value':..., 'fill':..., 'font':...}, ...], ...]"""
    wb = load_workbook(wb_path, data_only=True)
    try:
        ws = wb[sheet_name]
        return [[_cell_info(c) for c in row] for row in ws.iter_rows()]
    finally:
        wb.close()


def _read_first_sheet_rows_with_style(wb_path: str) -> List[List[dict]]:
    """读取第一个 sheet，返回带样式的单元格信息"""
    wb = load_workbook(wb_path, data_only=True)
    try:
        ws = wb.active
        return [[_cell_info(c) for c in row] for row in ws.iter_rows()]
    finally:
        wb.close()



#  写入函数升级：支持纯值或 dict 样式
def _copy_rows_to_ws(ws, rows: List[List], start_row: int = 1) -> int:
    """将行数据写入 worksheet，支持纯值或 dict 样式，返回最后写入的行号"""
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, cell_data in enumerate(row, start=1):
            if isinstance(cell_data, dict):
                val = cell_data.get("value")
                new_cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if cell_data.get("fill"):
                    new_cell.fill = cell_data["fill"]
                if cell_data.get("font"):
                    new_cell.font = cell_data["font"]
            else:
                ws.cell(row=r_idx, column=c_idx, value=cell_data)
    return start_row + len(rows) - 1 if rows else start_row


def _apply_header_style(ws, row: int, cols: int):
    """给指定行应用表头样式（蓝色背景 + 白字）"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_source_title(ws, row: int, cols: int, title: str):
    """插入数据源标题行并合并单元格"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(cols, 1))
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=12, color="1F2937")
    cell.fill = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")