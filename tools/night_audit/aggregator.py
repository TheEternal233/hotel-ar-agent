"""night_audit 汇总器：将多源对账结果合并为单一 Excel"""

import os
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font,  Alignment
from openpyxl.utils import get_column_letter

from .core import (
    _read_sheet_rows_with_style,
    _read_first_sheet_rows_with_style,
    _copy_rows_to_ws,
    _apply_header_style,
    _apply_source_title,
)


OTA_SHEET_NAME = "对账汇总"
GAP_ROWS = 5
DEFAULT_OTA_OUTPUT_DIR = ("output", "OTA对账")
DEFAULT_CARD_OUTPUT_DIR = ("output", "信用卡审核")


def _collect_ota_data(ota_paths: List[str]) -> List[tuple]:
    results = []
    for p in ota_paths:
        if not os.path.exists(p):
            continue
        rows = _read_sheet_rows_with_style(p, OTA_SHEET_NAME)
        if rows:
            results.append((Path(p).name, rows))
    return results


def _collect_card_data(card_paths: List[str]) -> List[tuple]:
    results = []
    for p in card_paths:
        if not os.path.exists(p):
            continue
        rows = _read_first_sheet_rows_with_style(p)
        if rows:
            results.append((Path(p).name, rows))
    return results


def _write_block(ws, start_row: int, title: str, blocks: List[tuple], inter_file_gap: int = 3) -> int:
    if not blocks:
        return start_row

    max_cols = max(
        max((len(r) for r in rows), default=0)
        for _, rows in blocks
    )

    _apply_source_title(ws, start_row, max_cols, title)
    current_row = start_row + 1

    for idx, (file_name, rows) in enumerate(blocks):
        # 文件小标题
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=max_cols)
        cell = ws.cell(row=current_row, column=1, value=f"来源: {file_name}")
        cell.font = Font(bold=True, size=11, color="374151")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

        # 数据行（带样式复制）
        _copy_rows_to_ws(ws, rows, current_row)

        # 表头样式（只覆盖第一行，数据行颜色保留）
        if rows:
            _apply_header_style(ws, current_row, max_cols)

        current_row += len(rows)

        # 文件之间空行
        if idx < len(blocks) - 1:
            current_row += inter_file_gap

    return current_row + GAP_ROWS


def aggregate_daily_check(
    ota_paths: List[str],
    card_paths: List[str],
    output_path: str = "",
) -> str:
    ota_blocks = _collect_ota_data(ota_paths)
    card_blocks = _collect_card_data(card_paths)

    if not ota_blocks and not card_blocks:
        raise ValueError("未找到有效的对账数据文件")

    if not output_path:
        base = Path(__file__).resolve().parent.parent.parent / "output"
        base.mkdir(parents=True, exist_ok=True)
        output_path = str(base / f"夜审汇总_{datetime.now().strftime('%Y%m%d')}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "夜审汇总"

    current_row = 1
    if ota_blocks:
        current_row = _write_block(ws, current_row, "OTA 对账结果", ota_blocks)
    if card_blocks:
        current_row = _write_block(ws, current_row, "信用卡对账结果", card_blocks)

    # 自动调整列宽
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if hasattr(cell, 'value') and cell.value is not None:
                try:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
                except Exception:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_path)
    wb.close()
    return output_path


def auto_aggregate_daily_check() -> str:
    """自动扫描 output/OTA对账/ 和 output/信用卡审核/ 目录，汇总生成夜审报告"""
    base = Path(__file__).resolve().parent.parent.parent
    ota_dir = base.joinpath(*DEFAULT_OTA_OUTPUT_DIR)
    card_dir = base.joinpath(*DEFAULT_CARD_OUTPUT_DIR)

    ota_paths = []
    if ota_dir.exists():
        ota_paths = sorted(
            [str(f) for f in ota_dir.iterdir() if f.is_file() and f.suffix in (".xlsx", ".xls")],
            key=lambda p: os.path.getmtime(p), reverse=True
        )

    card_paths = []
    if card_dir.exists():
        card_paths = sorted(
            [str(f) for f in card_dir.iterdir() if f.is_file() and f.suffix in (".xlsx", ".xls")],
            key=lambda p: os.path.getmtime(p), reverse=True
        )

    if not ota_paths and not card_paths:
        raise ValueError("未找到对账结果文件，请先执行 OTA对账 和 信用卡对账")

    return aggregate_daily_check(ota_paths, card_paths)