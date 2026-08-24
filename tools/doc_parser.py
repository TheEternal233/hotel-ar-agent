"""文档解析标准接口 —— 统一 Excel 读取、列映射、校验逻辑

所有工具模块通过此标准接口读取 xlsx，不再各自实现 ad-hoc 解析。

接口一览:
    read_sheet(path, header_row, sheet_name) -> (headers, [row_dicts])
    read_indexed(path, key, header_row, sheet_name) -> (headers, {key: row})
    read_mapped(path, mapping, header_row, sheet_name, cast) -> [standardized_dicts]
    read_rezen(path, header_row, sheet_name) -> [standardized_dicts]
    get_info(path, sheet_name) -> {sheet, cols, rows, headers}
    validate(path, required, sheet_name) -> (bool, msg)
"""

import os
import zipfile
from datetime import datetime
import openpyxl
import xml.etree.ElementTree as ET

from utils.file_cache import file_cache


def _open(path, sheet_name=None):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        return wb, ws
    except Exception:
        wb.close()
        raise


def _get_headers(ws, header_row=1):
    row_vals = [c.value for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    return [str(v) if v is not None else "" for v in row_vals]


def _parse_date(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, str) and val.strip():
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"]:
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                pass
    return None


@file_cache
def read_sheet(path, header_row=1, sheet_name=None, skip_empty=True, start_row=None, wb=None):
    close_wb = False
    if wb is None:
        wb, ws = _open(path, sheet_name)
        close_wb = True
    else:
        ws = wb[sheet_name] if sheet_name else wb.active
    try:
        headers = _get_headers(ws, header_row)
        start = start_row or (header_row + 1)
        rows = []
        for row in ws.iter_rows(min_row=start, values_only=True):
            record = dict(zip(headers, row))
            if skip_empty and all(v is None for v in record.values()):
                continue
            rows.append(record)
        return headers, rows
    finally:
        if close_wb:
            wb.close()


def read_indexed(path, key_column, header_row=1, sheet_name=None):
    wb, ws = _open(path, sheet_name)
    try:
        headers = _get_headers(ws, header_row)
        if key_column not in headers:
            raise KeyError(f"Column '{key_column}' not found. Available: {headers}")
        data = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            record = dict(zip(headers, row))
            key = record.get(key_column)
            if key is not None:
                data[str(key)] = record
        return headers, data
    finally:
        wb.close()


def read_mapped(path, column_map, header_row=1, sheet_name=None, cast=None):
    wb, ws = _open(path, sheet_name)
    try:
        headers = _get_headers(ws, header_row)
        headers_lower = [h.lower() for h in headers]

        field_to_col = {}
        for std_name, keywords in column_map.items():
            for kw in keywords:
                kw_lower = kw.lower()
                for idx, hl in enumerate(headers_lower):
                    if kw_lower in hl:
                        field_to_col[std_name] = idx
                        break
                if std_name in field_to_col:
                    break

        cast_fn = cast or {}
        records = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            record = {}
            for std_name, col_idx in field_to_col.items():
                val = row[col_idx] if col_idx < len(row) else None
                fn = cast_fn.get(std_name)
                if fn and val is not None:
                    try:
                        val = fn(val)
                    except (ValueError, TypeError):
                        val = None
                record[std_name] = val
            if any(v is not None for v in record.values()):
                records.append(record)
        return records
    finally:
        wb.close()


@file_cache
def read_rezen(path, header_row=1, sheet_name=None, wb=None):
    close_wb = False
    if wb is None:
        wb, ws = _open(path, sheet_name)
        close_wb = True
    else:
        ws = wb[sheet_name] if sheet_name else wb.active
    try:
        headers = _get_headers(ws, header_row)
        headers_lower = [h.lower() for h in headers]

        def _find_idx(*keywords):
            for kw in keywords:
                kwl = kw.lower()
                for i, h in enumerate(headers_lower):
                    if kwl in h:
                        return i
            return -1
        def _find_idx_exact(*keywords):
            """精确匹配：表头必须完全等于关键词，避免子串误匹配"""
            for kw in keywords:
                kwl = kw.lower()
                for i, h in enumerate(headers_lower):
                    if h == kwl:
                        return i
            return -1
        idx_bill = _find_idx("账单号")
        idx_type = _find_idx("类型")
        idx_date = _find_idx("日期")
        idx_time = _find_idx("结账时间")
        idx_room = _find_idx("房号")
        idx_amount = _find_idx("金额")
        idx_debit = _find_idx("借方")
        idx_credit = _find_idx("贷方")
        idx_written = _find_idx("已核销")
        idx_balance = _find_idx("余额")
        idx_note = _find_idx("财务备注", "备注")
        idx_bill_no = _find_idx("结账单号")
        idx_ext_order = _find_idx("外部订单号")
        idx_order = _find_idx_exact("订单号")
        idx_corp = _find_idx("协议单位")
        idx_name = _find_idx("姓名", "描述")
        idx_remark = _find_idx("转账注释", "订单备注")
        idx_operator = _find_idx("入账操作员")

        def _val(row, idx):
            if idx >= 0 and idx < len(row):
                return row[idx]
            return None

        records = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if all(v is None for v in row):
                continue
            rec = {
                "bill_id": str(_val(row, idx_bill) or ""),
                "type": str(_val(row, idx_type) or ""),
                "date": _parse_date(_val(row, idx_date)),
                "time": str(_val(row, idx_time) or ""),
                "room": str(_val(row, idx_room) or ""),
                "amount": float(_val(row, idx_amount) or 0),
                "debit": float(_val(row, idx_debit) or 0),
                "credit": float(_val(row, idx_credit) or 0),
                "written_off": float(_val(row, idx_written) or 0),
                "balance": float(_val(row, idx_balance) or 0),
                "note": str(_val(row, idx_note) or ""),
                "bill_no": str(_val(row, idx_bill_no) or ""),
                "ext_order": str(_val(row, idx_ext_order) or ""),
                "order": str(_val(row, idx_order) or ""),
                "corp": str(_val(row, idx_corp) or ""),
                "name": str(_val(row, idx_name) or ""),
                "remark": str(_val(row, idx_remark) or ""),
                "operator": str(_val(row, idx_operator) or ""),
            }
            records.append(rec)
        return records
    finally:
        if close_wb:
            wb.close()


@file_cache
def get_info(path, sheet_name=None):
    wb, ws = _open(path, sheet_name)
    try:
        headers = _get_headers(ws, 1)
        rows = sum(1 for _ in ws.iter_rows(min_row=2))
        return {
            "path": path,
            "filename": os.path.basename(path),
            "sheet": ws.title,
            "cols": len(headers),
            "rows": rows,
            "headers": headers,
        }
    finally:
        wb.close()


def validate(path, required_columns, sheet_name=None):
    wb = None
    try:
        wb, ws = _open(path, sheet_name)
        headers = _get_headers(ws, 1)
        headers_str = [str(h) for h in headers]
        missing = [c for c in required_columns if c not in headers_str]
        if missing:
            return False, f"Missing columns: {missing} (found: {headers_str})"
        return True, f"OK: {len(headers_str)} cols, all required present"
    except FileNotFoundError as e:
        return False, str(e)
    except Exception as e:
        return False, f"Error: {e}"
    finally:
        if wb is not None:
            wb.close()


# ===== 渠道列映射 —— 清远酒店真实数据 =====

OTA_CHANNEL_MAPPINGS = {
    "携程客房": {
        "header_row": 2,
        "columns": {
            "order_id": ["订单号"],
            "confirm_id": ["确认号"],
            "room_type": ["房型名称"],
            "checkin": ["入住日期"],
            "checkout": ["离店日期"],
            "guest_name": ["客人姓名"],
            "nights": ["间夜"],
            "settle_amount": ["结算价"],
            "currency": ["币种"],
            "base_amount": ["订单折前底价/卖价"],
            "discount": ["底价/卖价折扣金额"],
        },
        "order_id_col": "order_id",
        "amount_col": "settle_amount",
    },
    "携程餐饮": {
        "header_row": 1,
        "columns": {
            "voucher_no": ["券号"],
            "food_voucher": ["美食林券号"],
            "use_time": ["验券时间"],
            "product_id": ["产品ID"],
            "product_name": ["产品名称"],
            "base_price": ["底价"],
            "sell_price": ["卖价"],
            "currency": ["币种"],
            "buy_time": ["购买时间"],
        },
        "order_id_col": "voucher_no",
        "amount_col": "sell_price",
    },
    "美团客房": {
        "header_row": 1,
        "columns": {
            "order_id": ["美团订单号"],
            "hotel_name": ["酒店名称"],
            "city": ["城市名称"],
            "checkin": ["入住日期"],
            "checkout": ["离店日期"],
            "room_type": ["房型名称"],
            "guest_name": ["入住人姓名"],
            "nights": ["间夜"],
            "amount": ["结算金额","美团结算价"],
            "order_status": ["订单状态"],
        },
        "order_id_col": "order_id",
        "amount_col": "amount",
    },
    "美团餐饮": {
        "header_row": 1,
        "columns": {
            "voucher_no": ["券号"],
            "order_id": ["订单号"],
            "revenue_time": ["收益时间"],
            "order_time": ["下单时间"],
            "consume_time": ["消费时间"],
            "sell_price": ["售价（美团售价）"],
            "product_name": ["产品名称"],
        },
        "order_id_col": "order_id",
        "amount_col": "sell_price",
    },
    "飞猪": {
        "header_row": 1,
        "columns": {
            "order_id": ["订单号", "套餐订单号"],
            "calendar_order_id": ["日历房订单号"],
            "guest_name": ["入住人"],
            "checkin": ["入住日期"],
            "checkout": ["离店日期"],
            "room_type": ["房型"],
            "nights": ["间夜"],
            "amount": ["分账金额","分账总金额","结算价", "卖家实收"],
        },
        "order_id_col": "order_id",
        "amount_col": "amount",
    },
    "抖音": {
        "header_row": 1,
        "columns": {
            "verify_time": ["核销时间"],
            "order_id": ["订单编号"],
            "linked_order": ["关联单号"],
            "voucher_code": ["券码"],
            "verify_id": ["核销ID"],
            "amount": ["订单实收金额","核销金额", "金额"],
        },
        "order_id_col": "order_id",
        "amount_col": "amount",
    },
    "向蜜鸟": {
        "header_row": 1,
        "columns": {
            "serial_no": ["流水号"],
            "biz_line": ["业务线"],
            "order_id": ["订单号"],
            "identify_no": ["识别号"],
            "fee_detail": ["费用明细1"],
            "amount": ["金额"],
        },
        "order_id_col": "order_id",
        "amount_col": "amount",
    },
}

AGING_MAPPING = {
    "customer": ["客户", "customer", "客户名称", "name"],
    "due_date": ["到期", "due", "到期日", "due_date", "截止"],
    "amount":   ["金额", "amount", "应收金额", "balance"],
}

CTRIP_MAPPING = {
    "order_id":   ["order", "id", "订单号", "order_id"],
    "nights":     ["night", "间夜", "nights"],
    "amount":     ["金额", "amount", "房费"],
    "commission": ["comm", "佣金", "commission", "平台佣金"],
}

CARD_MAPPING = {
    "date":   ["date", "日期", "交易日期", "trade_date"],
    "amount": ["amount", "金额", "交易金额", "trade_amount"],
    "card":   ["card", "卡号", "card_number", "card_no"],
}

@file_cache
def read_ota_channel(path, channel_name, sheet_name=None, wb=None):
    cfg = OTA_CHANNEL_MAPPINGS.get(channel_name)
    if not cfg:
        raise ValueError(f"Unknown channel: {channel_name}. Available: {list(OTA_CHANNEL_MAPPINGS.keys())}")

    close_wb = False
    if wb is None:
        wb, ws = _open(path, sheet_name)
        close_wb = True
    else:
        ws = wb[sheet_name] if sheet_name else wb.active
    try:
        hr = cfg["header_row"]
        headers = _get_headers(ws, hr)
        headers_lower = [h.lower() for h in headers]

        col_map = cfg["columns"]
        field_to_col = {}
        for std_name, keywords in col_map.items():
            for kw in keywords:
                kwl = kw.lower()
                for idx, hl in enumerate(headers_lower):
                    if kwl in hl:
                        field_to_col[std_name] = idx
                        break
                if std_name in field_to_col:
                    break

        records = []
        start_row = hr + 1
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if all(v is None for v in row):
                continue
            rec = {"channel": channel_name}
            for std_name, col_idx in field_to_col.items():
                val = row[col_idx] if col_idx < len(row) else None
                rec[std_name] = val
            records.append(rec)

        return records
    finally:
        if close_wb:
            wb.close()


@file_cache
def detect_ota_channel(path):
    wb, ws = _open(path)
    try:
        headers = _get_headers(ws, 1)
        headers_str = " ".join(str(h) for h in headers if h)

        rezen_markers = ["账单号", "外部订单号", "协议单位"]
        rezen_score = sum(1 for m in rezen_markers if m in headers_str)
        if rezen_score >= 2:
            return "rezen"

        if "财务总对账" in wb.sheetnames and "PMS" in wb.sheetnames:
            return "向蜜鸟"

        if "美团订单号" in headers_str:
            return "美团客房"
        if "券号" in headers_str and "美食林券号" in headers_str:
            return "携程餐饮"
        if "券号" in headers_str and "订单号" in headers_str and "售价" in headers_str:
            return "美团餐饮"
        if "核销时间" in headers_str and "订单编号" in headers_str:
            return "抖音"
        if "套餐订单号" in headers_str or ("订单号" in headers_str and "入住人" in headers_str):
            return "飞猪"

        # Check for 携程客房 (header in row 2)
        if "预付订单明细" in headers_str or "订单类型" in headers_str:
            row2 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
            row2_str = " ".join(str(v) for v in row2 if v)
            if "订单号" in row2_str and "结算价" in row2_str:
                return "携程客房"

        return None
    finally:
        wb.close()

def _fast_read_xlsx_headers(path, max_rows=2):
    """用 zipfile+xml 快速读取 xlsx 的前几行表头，不经过 openpyxl 完整加载"""
    ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
          'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
    shared_strings = []
    sheet_names = []

    try:
        with zipfile.ZipFile(path, 'r') as zf:
            # 读取共享字符串表
            if 'xl/sharedStrings.xml' in zf.namelist():
                with zf.open('xl/sharedStrings.xml') as f:
                    tree = ET.parse(f)
                    for si in tree.findall('.//main:t', ns):
                        shared_strings.append(si.text or '')

            # 读取工作簿关系，获取 sheet 名称
            if 'xl/workbook.xml' in zf.namelist():
                with zf.open('xl/workbook.xml') as f:
                    tree = ET.parse(f)
                    for sheet in tree.findall('.//main:sheet', ns):
                        sheet_names.append(sheet.get('name', ''))

            # 读取第一个工作表的前几行
            sheet_path = 'xl/worksheets/sheet1.xml'
            if sheet_path not in zf.namelist():
                return [], sheet_names

            with zf.open(sheet_path) as f:
                tree = ET.parse(f)
                rows = []
                for row in tree.findall('.//main:row', ns)[:max_rows]:
                    cells = []
                    for cell in row.findall('.//main:c', ns):
                        val = ''
                        cell_type = cell.get('t', '')
                        v_elem = cell.find('main:v', ns)
                        if v_elem is not None and v_elem.text:
                            if cell_type == 's':
                                idx = int(v_elem.text)
                                if idx < len(shared_strings):
                                    val = shared_strings[idx]
                            else:
                                val = v_elem.text
                        cells.append(val)
                    rows.append(cells)
                return rows, sheet_names
    except Exception:
        return [], sheet_names

def _detect_ota_channel_from_headers(headers, sheet_names, rows):
    """根据表头内容检测渠道"""
    headers_str = " ".join(str(h) for h in headers if h)

    # 优先检测 OTA 渠道（向蜜鸟有 rezen 标记，优先判断）
    if "财务总对账" in sheet_names and "PMS" in sheet_names:
        return "向蜜鸟"

    if "美团订单号" in headers_str:
        return "美团客房"
    if "券号" in headers_str and "美食林券号" in headers_str:
        return "携程餐饮"
    if "券号" in headers_str and "订单号" in headers_str and "售价" in headers_str:
        return "美团餐饮"
    if "核销时间" in headers_str and "订单编号" in headers_str:
        return "抖音"
    if "套餐订单号" in headers_str or ("订单号" in headers_str and "入住人" in headers_str):
        return "飞猪"

    # Check for 携程客房 (header in row 2)
    if "预付订单明细" in headers_str or "订单类型" in headers_str:
        if len(rows) > 1:
            row2_str = " ".join(str(v) for v in rows[1] if v)
            if "订单号" in row2_str and "结算价" in row2_str:
                return "携程客房"

    # 最后检测 rezen（PMS 文件）
    rezen_markers = ["账单号", "外部订单号", "协议单位"]
    rezen_score = sum(1 for m in rezen_markers if m in headers_str)
    if rezen_score >= 2:
        return "rezen"

    return None


def detect_ota_channel_fast(path):
    """快速检测 OTA 渠道，不打开 openpyxl（用于批量对账前置判断）

    对于包含合并单元格等特殊格式的文件，会回退到 openpyxl 读取。
    """
    # 先尝试用 zipfile+xml 快速读取
    rows, sheet_names = _fast_read_xlsx_headers(path, max_rows=3)
    if rows:
        headers = [str(h) for h in rows[0] if h]
        # 如果快速读取到了有效表头，直接使用
        if headers:
            result = _detect_ota_channel_from_headers(headers, sheet_names, rows)
            if result:
                return result

    # 快速读取失败或结果为空，回退到 openpyxl 完整读取
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            ws = wb.active
            sheet_names = [s.title for s in wb.worksheets]
            # 尝试读取前3行，找到非空的表头行
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                rows.append([str(v) if v is not None else "" for v in row])
            for row in rows:
                headers = [h for h in row if h]
                if headers:
                    result = _detect_ota_channel_from_headers(headers, sheet_names, rows)
                    if result:
                        return result
        finally:
            wb.close()
    except Exception:
        pass

    return None