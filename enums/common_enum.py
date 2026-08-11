import os

from openpyxl.styles import Border, PatternFill, Font, Side
from enums.paths import BASE_DIR, OUTPUT_DIR as OUT_DIR

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")