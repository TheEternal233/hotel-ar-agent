
from copy import copy

import openpyxl

from tools.doc_parser import _parse_date
from utils.file_cache import file_cache


def _unique_sheet_name(wb, name):
    """生成不重复的工作表名称。"""
    if name not in wb.sheetnames:
        return name
    i = 1
    while True:
        new_name = f"{name}_{i}"
        if new_name not in wb.sheetnames:
            return new_name
        i += 1


def _copy_sheet_to_wb(src_ws, dst_wb, title=None):
    """将源工作表复制到目标工作簿，保留合并单元格、列宽、行高、冻结窗格、单元格内容和样式。"""
    if title is None:
        title = src_ws.title
    title = _unique_sheet_name(dst_wb, title)
    dst_ws = dst_wb.create_sheet(title=title)

    # 合并单元格
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    # 列宽
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width

    # 行高
    for row, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row].height = dim.height

    # 冻结窗格
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes

    # 单元格内容、公式与样式
    # 优化：直接遍历 _cells 字典，只处理实际存在的单元格，跳过空单元格
    # 这比 iter_rows() 遍历整个矩形区域更高效，尤其是稀疏表格
    for (row, col), cell in src_ws._cells.items():
        new_cell = dst_ws.cell(row=row, column=col, value=cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
        if cell.hyperlink:
            new_cell.hyperlink = copy(cell.hyperlink)
        if cell.comment:
            new_cell.comment = copy(cell.comment)

    return dst_ws



@file_cache
def read_xiangminiao(path, wb=None):
    """读取向蜜鸟对账文件（同文件多sheet）

    使用 read_only=True 减少内存占用，因为该函数只读取数据不修改。
    对账文件中的字段通常为原始数值，不涉及公式计算。

    支持传入已加载的 workbook 对象(wb)，避免重复从磁盘加载。
    调用方负责关闭传入的 workbook。
    """
    close_wb = False
    if wb is None:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        close_wb = True
    try:
        def _find_idx(headers_lower, *keywords):
            for kw in keywords:
                kwl = kw.lower()
                for i, h in enumerate(headers_lower):
                    if kwl in h:
                        return i
            return -1

        def _val(row, idx):
            if idx >= 0 and idx < len(row):
                return row[idx]
            return None

        # ---------- 1. 读取 PMS sheet ----------
        ws_pms = wb["PMS"]
        pms_headers = [str(v) if v is not None else "" for v in next(ws_pms.iter_rows(min_row=1, max_row=1, values_only=True))]
        pms_headers_lower = [h.lower() for h in pms_headers]

        pms_idx_bill = _find_idx(pms_headers_lower, "账单号")
        pms_idx_type = _find_idx(pms_headers_lower, "类型")
        pms_idx_date = _find_idx(pms_headers_lower, "日期")
        pms_idx_time = _find_idx(pms_headers_lower, "结账时间")
        pms_idx_room = _find_idx(pms_headers_lower, "房号")
        pms_idx_amount = _find_idx(pms_headers_lower, "金额")
        pms_idx_debit = _find_idx(pms_headers_lower, "借方")
        pms_idx_credit = _find_idx(pms_headers_lower, "贷方")
        pms_idx_written = _find_idx(pms_headers_lower, "已核销")
        pms_idx_balance = _find_idx(pms_headers_lower, "余额")
        pms_idx_note = _find_idx(pms_headers_lower, "财务备注", "备注")
        pms_idx_bill_no = _find_idx(pms_headers_lower, "结账单号")
        pms_idx_ext_order = _find_idx(pms_headers_lower, "外部订单号")
        pms_idx_order = _find_idx(pms_headers_lower, "订单号")
        pms_idx_remark = _find_idx(pms_headers_lower, "转账注释", "订单备注")
        pms_idx_operator = _find_idx(pms_headers_lower, "入账操作员")
        pms_idx_name = _find_idx(pms_headers_lower, "姓名", "描述")

        rezen_records = []
        for row in ws_pms.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rec = {
                "bill_id": str(_val(row, pms_idx_bill) or ""),
                "type": str(_val(row, pms_idx_type) or ""),
                "date": _parse_date(_val(row, pms_idx_date)),
                "time": str(_val(row, pms_idx_time) or ""),
                "room": str(_val(row, pms_idx_room) or ""),
                "amount": float(_val(row, pms_idx_amount) or 0),
                "debit": float(_val(row, pms_idx_debit) or 0),
                "credit": float(_val(row, pms_idx_credit) or 0),
                "written_off": float(_val(row, pms_idx_written) or 0),
                "balance": float(_val(row, pms_idx_balance) or 0),
                "note": str(_val(row, pms_idx_note) or ""),
                "bill_no": str(_val(row, pms_idx_bill_no) or ""),
                "ext_order": str(_val(row, pms_idx_ext_order) or ""),
                "order": str(_val(row, pms_idx_order) or ""),
                "remark": str(_val(row, pms_idx_remark) or ""),
                "operator": str(_val(row, pms_idx_operator) or ""),
                "name": str(_val(row, pms_idx_name) or ""),
            }
            rezen_records.append(rec)

        # ---------- 2. 读取 财务总对账 sheet -> OTA记录 ----------
        ws_ota = wb["财务总对账"]
        ota_headers = [str(v) if v is not None else "" for v in next(ws_ota.iter_rows(min_row=1, max_row=1, values_only=True))]
        ota_headers_lower = [h.lower() for h in ota_headers]

        ota_idx_order = _find_idx(ota_headers_lower, "订单号")
        ota_idx_biz_type = _find_idx(ota_headers_lower, "业务类型")
        ota_idx_identify = _find_idx(ota_headers_lower, "识别号")
        ota_idx_fee1 = _find_idx(ota_headers_lower, "费用明细1")
        ota_idx_pay_type = _find_idx(ota_headers_lower, "支付方式")
        ota_idx_status = _find_idx(ota_headers_lower, "状态")
        ota_idx_order_amount = _find_idx(ota_headers_lower, "订单金额")
        ota_idx_card_pay = _find_idx(ota_headers_lower, "储值卡支付金额")
        ota_idx_settle = _find_idx(ota_headers_lower, "结算金额")
        ota_idx_actual = _find_idx(ota_headers_lower, "实际结算")

        ota_records = []
        for row in ws_ota.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            order_id = str(_val(row, ota_idx_order) or "")
            identify_no = str(_val(row, ota_idx_identify) or "")
            settle_amount=float(_val(row, ota_idx_settle) or 0)
            order_amount=float(_val(row, ota_idx_order_amount) or 0)
            if order_id == "总计" :
                continue

            if not order_id and not identify_no and settle_amount==0 and order_amount==0:
                continue
            rec = {
                "channel": "向蜜鸟",
                "order_id": order_id,
                "biz_type": str(_val(row, ota_idx_biz_type) or ""),
                "identify_no": str(_val(row, ota_idx_identify) or ""),
                "fee_detail": str(_val(row, ota_idx_fee1) or ""),
                "pay_type": str(_val(row, ota_idx_pay_type) or ""),
                "status": str(_val(row, ota_idx_status) or ""),
                "order_amount": float(_val(row, ota_idx_order_amount) or 0),
                "card_pay_amount": float(_val(row, ota_idx_card_pay) or 0),
                "settle_amount": float(_val(row, ota_idx_settle) or 0),
                "actual_settle": float(_val(row, ota_idx_actual) or 0),
            }
            ota_records.append(rec)

        # ---------- 3. 读取 储值卡消费对账 sheet ----------
        ws_card = wb["储值卡消费对账"]
        card_headers = [str(v) if v is not None else "" for v in next(ws_card.iter_rows(min_row=1, max_row=1, values_only=True))]
        card_headers_lower = [h.lower() for h in card_headers]

        card_idx_order = _find_idx(card_headers_lower, "订单号")
        card_idx_amount = _find_idx(card_headers_lower, "储值卡消费金额")
        card_idx_pay_type = _find_idx(card_headers_lower, "支付方式")
        card_idx_status = _find_idx(card_headers_lower, "消费状态")

        card_records = []
        for row in ws_card.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            order_id = str(_val(row, card_idx_order) or "")
            if order_id == "总计" or not order_id:
                continue
            rec = {
                "order_id": order_id,
                "card_amount": float(_val(row, card_idx_amount) or 0),
                "pay_type": str(_val(row, card_idx_pay_type) or ""),
                "status": str(_val(row, card_idx_status) or ""),
            }
            card_records.append(rec)

        return ota_records, card_records, rezen_records
    finally:
        if close_wb:
            wb.close()